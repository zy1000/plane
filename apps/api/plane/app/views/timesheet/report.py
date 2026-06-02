import datetime
import io
from urllib.parse import quote

from django.db.models import Q
from django.http import FileResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from plane.app.serializers.timesheet import TimeSheetReportListSerializer
from plane.app.views import BaseViewSet
from plane.db.models import TimeSheet


def _parse_ids(raw: str):
    """解析 ?ids=a,b,c 为 uuid 字符串列表，过滤空值。"""
    if not raw:
        return []
    return [item.strip() for item in raw.split(",") if item.strip()]


class TimeSheetReportViewSet(BaseViewSet):
    model = TimeSheet
    serializer_class = TimeSheetReportListSerializer

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related(
                "project",
                "issue",
                "test_case",
                "category",
                "member",
                "member__extra_info",
            )
            .exclude()
        )

    def _apply_filters(self, request, slug):
        """按查询参数构造 queryset，供 list / export 复用。

        project_id / member_id / category_id / category_key 均支持单值或逗号分隔的多值。
        """
        params = request.query_params
        query = self.get_queryset()

        if not params.get("all_workspace"):
            query = query.filter(project__workspace__slug=slug)

        if project_ids := _parse_ids(params.get("project_id", "")):
            query = query.filter(project_id__in=project_ids)

        # 项目编号筛选：特殊值 "__empty__" 代表项目编号为空（null 或空串）
        if pms_names := _parse_ids(params.get("pms_project_name", "")):
            empty_sentinel = "__empty__"
            real_names = [name for name in pms_names if name != empty_sentinel]
            pms_query = Q()
            if empty_sentinel in pms_names:
                pms_query |= Q(project__pms_project_name__isnull=True) | Q(
                    project__pms_project_name=""
                )
            if real_names:
                pms_query |= Q(project__pms_project_name__in=real_names)
            if pms_query:
                query = query.filter(pms_query)

        if start_time := params.get("start_time"):
            query = query.filter(date__gte=start_time)

        if end_time := params.get("end_time"):
            query = query.filter(date__lte=end_time)

        if member_ids := _parse_ids(params.get("member_id", "")):
            query = query.filter(member_id__in=member_ids)

        if category_ids := _parse_ids(params.get("category_id", "")):
            query = query.filter(category_id__in=category_ids)

        category_key_raw = (
            params.get("category_key") or params.get("category__key") or ""
        )
        if category_keys := _parse_ids(category_key_raw):
            query = query.filter(category__key__in=category_keys)

        return query.order_by("-date", "-start_time", "-id")

    def list(self, request, slug):
        queryset = self._apply_filters(request, slug)
        return self.paginate(
            order_by="-date",
            request=request,
            queryset=queryset,
            default_per_page=50,
            max_per_page=200,
            on_results=lambda result: TimeSheetReportListSerializer(
                result, many=True
            ).data,
        )

    @action(detail=False, methods=["get"], url_path="export")
    def export_xlsx(self, request, slug):
        """导出 xlsx。若提供 ids 则仅导出勾选项，否则导出当前过滤全量。"""
        queryset = self._apply_filters(request, slug)
        ids = _parse_ids(request.query_params.get("ids", ""))
        if ids:
            queryset = queryset.filter(id__in=ids)

        columns = [
            ("项目编号", "pms_project_name"),
            ("项目名称", "project_name"),
            ("工作项", "issue_name"),
            ("测试用例", "case_name"),
            ("成员", "member_name"),
            ("工号", "employee_id"),
            ("部门", "department"),
            ("日期", "date"),
            ("开始时间", "start_time"),
            ("结束时间", "end_time"),
            ("工时", "hours"),
            ("类别", "category_name"),
            ("描述", "description"),
        ]

        workbook = Workbook()
        ws = workbook.active
        ws.title = "工时报表"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_align = Alignment(horizontal="center", vertical="center")

        ws.append([label for label, _ in columns])
        for col_idx, _ in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        def format_value(record, key):
            if key == "pms_project_name":
                return record.project.pms_project_name or ""
            if key == "project_name":
                return record.project.name or ""
            if key == "issue_name":
                return record.issue.name if record.issue_id else ""
            if key == "case_name":
                return record.test_case.name if record.test_case_id else ""
            if key == "member_name":
                return (
                    getattr(record.member, "display_name", "")
                    or getattr(record.member, "email", "")
                    or ""
                )
            if key in ("employee_id", "department"):
                if not record.member_id:
                    return ""
                try:
                    extra_info = record.member.extra_info
                except Exception:
                    return ""
                return getattr(extra_info, key, "") or ""
            if key == "category_name":
                return record.category.name if record.category_id else ""
            if key == "date":
                return record.date.strftime("%Y-%m-%d") if record.date else ""
            if key in ("start_time", "end_time"):
                value = getattr(record, key, None)
                if not value:
                    return ""
                return value.strftime("%H:%M")
            if key == "hours":
                return str(record.hours) if record.hours is not None else ""
            if key == "description":
                return record.description or ""
            return ""

        for item in queryset.iterator():
            ws.append([format_value(item, key) for _, key in columns])

        widths = [16, 24, 30, 30, 14, 14, 20, 12, 10, 10, 8, 14, 40]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

        bio = io.BytesIO()
        workbook.save(bio)
        bio.seek(0)

        filename = f"timesheet-report-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        response = FileResponse(
            bio,
            as_attachment=True,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{quote(filename)}"
        )
        return response

    @staticmethod
    def _resolve_month_range(month_param):
        """把 YYYY-MM 解析为 [当月1号, 下月1号) 的左闭右开区间。

        month_param 为空时回落到当前月份。解析失败抛出 ValueError，由调用方转 400。
        """
        if month_param:
            year_str, month_str = month_param.split("-")
            first_day = datetime.date(int(year_str), int(month_str), 1)
        else:
            today = timezone.now().date()
            first_day = today.replace(day=1)

        if first_day.month == 12:
            next_month_first = datetime.date(first_day.year + 1, 1, 1)
        else:
            next_month_first = datetime.date(first_day.year, first_day.month + 1, 1)
        return first_day, next_month_first

    @staticmethod
    def _serialize_record(record):
        """将单条工时记录序列化为英文 key 的字典，与导出列保持一致。"""
        member = record.member
        employee_id = ""
        department = ""
        if record.member_id:
            try:
                extra_info = member.extra_info
            except Exception:
                extra_info = None
            if extra_info is not None:
                employee_id = getattr(extra_info, "employee_id", "") or ""
                department = getattr(extra_info, "department", "") or ""

        return {
            "pms_project_name": record.project.pms_project_name or "",
            "issue_name": record.issue.name if record.issue_id else "",
            "case_name": record.test_case.name if record.test_case_id else "",
            "member_name": (
                getattr(member, "display_name", "")
                or getattr(member, "email", "")
                or ""
            ),
            "employee_id": employee_id,
            "department": department,
            "date": record.date.strftime("%Y-%m-%d") if record.date else "",
            "start_time": (
                record.start_time.strftime("%H:%M") if record.start_time else ""
            ),
            "end_time": record.end_time.strftime("%H:%M") if record.end_time else "",
            "hours": str(record.hours) if record.hours is not None else "",
            "category": "项目工时",
            "description": record.description or "",
        }

    @action(detail=False, methods=["get"], url_path="export-json")
    def export_json(self, request):
        """公开接口（无需鉴权）：按月份与用户名导出用户填报的工时数据，返回 JSON。

        查询参数：
          - month: 形如 2026-06，控制导出哪个月的数据；不传默认当前月份。
          - user:  用户显示名（display_name），模糊匹配；不传默认全部用户。
        """
        month_param = (request.query_params.get("month") or "").strip()
        try:
            first_day, next_month_first = self._resolve_month_range(month_param)
        except (ValueError, TypeError):
            return Response(
                {"error": "month 参数格式不正确，应为 YYYY-MM，例如 2026-06。"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        queryset = self.get_queryset().filter(
            date__gte=first_day, date__lt=next_month_first
        )

        user_param = (request.query_params.get("user") or "").strip()
        if user_param:
            queryset = queryset.filter(member__display_name__icontains=user_param)

        queryset = queryset.order_by("-date", "-start_time", "-id")

        results = [self._serialize_record(item) for item in queryset.iterator()]

        return Response(
            {
                "month": first_day.strftime("%Y-%m"),
                "user": user_param or None,
                "count": len(results),
                "results": results,
            },
            status=status.HTTP_200_OK,
        )
