# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.

from collections import defaultdict
from datetime import datetime, timedelta
import io
from typing import Any, Dict, List, Optional
from urllib.parse import quote

from django.db.models import Prefetch
from django.http import FileResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import status
from rest_framework.response import Response

from plane.app.permissions import ROLE, allow_permission
from plane.app.views.base import BaseAPIView
from plane.db.models import (
    CycleOverdueRecord,
    Issue,
    IssueAssignee,
    ReleaseOverduePhase,
    ReleaseOverdueRecord,
    TestPlan,
)


class WorkspaceOverdueAnalyticsEndpoint(BaseAPIView):
    ALLOWED_STATUS = {"active", "all", "resolved"}
    ALLOWED_ENTITY_TYPES = {"issue", "cycle", "release", "test_plan"}
    ALLOWED_DATE_FIELDS = {"deadline", "overdue_since"}
    ENTITY_LABEL_MAP = {
        "issue": "工作项",
        "cycle": "迭代",
        "release": "发布",
        "test_plan": "测试计划",
    }

    @staticmethod
    def _parse_csv_ids(raw_value: Optional[str]) -> List[str]:
        if not raw_value:
            return []
        return [value.strip() for value in raw_value.split(",") if value and value.strip()]

    @staticmethod
    def _as_iso_date(value) -> Optional[str]:
        if value is None:
            return None
        if hasattr(value, "isoformat"):
            return value.isoformat()
        return None

    @staticmethod
    def _as_iso_datetime(value) -> Optional[str]:
        if value is None:
            return None
        if hasattr(value, "isoformat"):
            return value.isoformat()
        return None

    @staticmethod
    def _resolve_user_name(user) -> str:
        return user.display_name or f"{user.first_name or ''} {user.last_name or ''}".strip() or "-"

    @staticmethod
    def _resolve_issue_identifier(issue) -> str:
        project_identifier = getattr(issue.project, "identifier", None) or "-"
        return f"{project_identifier}-{issue.sequence_id}"

    @staticmethod
    def _calculate_overdue_days(
        overdue_since_date,
        *,
        is_active: bool,
        ended_at,
        today,
    ) -> int:
        if overdue_since_date is None:
            return 0

        end_date = today if is_active or ended_at is None else timezone.localdate(ended_at)
        return max((end_date - overdue_since_date).days, 0)

    def _extract_filters_from_request(self, request):
        status_filter = request.GET.get("status", "all")
        entity_type = request.GET.get("entity_type")
        project_ids = self._parse_csv_ids(request.GET.get("project_ids"))
        return status_filter, entity_type, project_ids

    def _validate_filters(self, *, status_filter: str, entity_type: Optional[str]):
        if status_filter not in self.ALLOWED_STATUS:
            return Response(
                {"error": "status 必须是 active、resolved 或 all"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if entity_type and entity_type not in self.ALLOWED_ENTITY_TYPES:
            return Response(
                {"error": "entity_type 必须是 issue、cycle、release 或 test_plan"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return None

    @staticmethod
    def _normalize_filter_date(raw_value: Optional[str]) -> Optional[str]:
        if not raw_value:
            return None

        value = raw_value.strip()[:10]
        try:
            datetime.strptime(value, "%Y-%m-%d")
        except ValueError:
            return None

        return value

    @staticmethod
    def _normalize_record_date(value: Optional[str]) -> Optional[str]:
        if not value:
            return None
        return str(value)[:10]

    @classmethod
    def _apply_date_filter(
        cls,
        records: List[Dict[str, Any]],
        *,
        date_field: str,
        start_date: Optional[str],
        end_date: Optional[str],
    ) -> List[Dict[str, Any]]:
        if not start_date and not end_date:
            return records

        filtered_records: List[Dict[str, Any]] = []
        for record in records:
            date_value = cls._normalize_record_date(record.get(date_field))
            if not date_value:
                continue
            if start_date and date_value < start_date:
                continue
            if end_date and date_value > end_date:
                continue
            filtered_records.append(record)

        return filtered_records

    def _build_issue_records(self, *, base_filters: Dict[str, Any], today, status_filter: str) -> List[Dict[str, Any]]:
        # Issue 没有独立历史逾期记录表，因此只返回当前仍逾期的数据
        if status_filter == "resolved":
            return []

        active_assignees_prefetch = Prefetch(
            "issue_assignee",
            queryset=IssueAssignee.objects.filter(deleted_at__isnull=True).select_related("assignee"),
            to_attr="active_issue_assignees",
        )
        overdue_issues = (
            Issue.issue_objects.filter(
                **base_filters,
                target_date__isnull=False,
                target_date__lt=today,
            )
            .exclude(state__group__in=["completed", "cancelled"])
            .select_related("project", "state")
            .prefetch_related(active_assignees_prefetch)
            .order_by("-target_date")
        )

        issue_records: List[Dict[str, Any]] = []
        for issue in overdue_issues:
            overdue_since_date = issue.target_date + timedelta(days=1) if issue.target_date else None
            assignees: List[Dict[str, Any]] = []
            seen_assignee_ids = set()
            for issue_assignee in getattr(issue, "active_issue_assignees", []):
                user = getattr(issue_assignee, "assignee", None)
                if user is None:
                    continue
                user_id = str(user.id)
                if user_id in seen_assignee_ids:
                    continue
                seen_assignee_ids.add(user_id)
                assignees.append(
                    {
                        "id": user_id,
                        "display_name": self._resolve_user_name(user),
                        "avatar_url": user.avatar_url or "",
                    }
                )

            issue_records.append(
                {
                    "entity_type": "issue",
                    "entity_id": str(issue.id),
                    "name": issue.name,
                    "identifier": self._resolve_issue_identifier(issue),
                    "project_id": str(issue.project_id),
                    "project_name": issue.project.name if issue.project else "-",
                    "deadline": self._as_iso_date(issue.target_date),
                    "overdue_since": self._as_iso_date(overdue_since_date),
                    "ended_at": None,
                    "is_active": True,
                    "overdue_days": self._calculate_overdue_days(
                        overdue_since_date,
                        is_active=True,
                        ended_at=None,
                        today=today,
                    ),
                    "phase": None,
                    "status_label": issue.state.name if issue.state else "-",
                    "assignees": assignees,
                }
            )

        return issue_records

    def _build_cycle_records(self, *, base_filters: Dict[str, Any], today, status_filter: str) -> List[Dict[str, Any]]:
        cycle_overdues = CycleOverdueRecord.objects.filter(
            **base_filters,
            deleted_at__isnull=True,
        ).select_related("cycle", "project")

        if status_filter == "active":
            cycle_overdues = cycle_overdues.filter(ended_at__isnull=True)
        elif status_filter == "resolved":
            cycle_overdues = cycle_overdues.filter(ended_at__isnull=False)

        cycle_overdues = cycle_overdues.order_by("-started_at")

        cycle_records: List[Dict[str, Any]] = []
        for record in cycle_overdues:
            overdue_since_date = timezone.localdate(record.started_at) if record.started_at else None
            is_active = record.ended_at is None
            deadline = None
            if record.cycle and record.cycle.end_date:
                deadline = record.cycle.end_date.date()

            cycle_records.append(
                {
                    "entity_type": "cycle",
                    "entity_id": str(record.cycle_id),
                    "name": record.cycle.name if record.cycle else "-",
                    "identifier": None,
                    "project_id": str(record.project_id),
                    "project_name": record.project.name if record.project else "-",
                    "deadline": self._as_iso_date(deadline),
                    "overdue_since": self._as_iso_date(overdue_since_date),
                    "ended_at": self._as_iso_datetime(record.ended_at),
                    "is_active": is_active,
                    "overdue_days": self._calculate_overdue_days(
                        overdue_since_date,
                        is_active=is_active,
                        ended_at=record.ended_at,
                        today=today,
                    ),
                    "phase": None,
                    "status_label": record.cycle.status if record.cycle else "-",
                    "assignees": [],
                }
            )

        return cycle_records

    def _build_release_records(self, *, base_filters: Dict[str, Any], today, status_filter: str) -> List[Dict[str, Any]]:
        release_overdues = ReleaseOverdueRecord.objects.filter(
            **base_filters,
            deleted_at__isnull=True,
        ).select_related("release", "project")

        if status_filter == "active":
            release_overdues = release_overdues.filter(ended_at__isnull=True)
        elif status_filter == "resolved":
            release_overdues = release_overdues.filter(ended_at__isnull=False)

        release_overdues = release_overdues.order_by("-started_at")

        release_records: List[Dict[str, Any]] = []
        for record in release_overdues:
            overdue_since_date = timezone.localdate(record.started_at) if record.started_at else None
            is_active = record.ended_at is None

            deadline = None
            if record.release:
                if record.phase == ReleaseOverduePhase.DEV:
                    deadline = record.release.test_handoff_date
                elif record.phase == ReleaseOverduePhase.TEST:
                    deadline = record.release.target_date

            release_records.append(
                {
                    "entity_type": "release",
                    "entity_id": str(record.release_id),
                    "name": record.release.name if record.release else "-",
                    "identifier": None,
                    "project_id": str(record.project_id),
                    "project_name": record.project.name if record.project else "-",
                    "deadline": self._as_iso_date(deadline),
                    "overdue_since": self._as_iso_date(overdue_since_date),
                    "ended_at": self._as_iso_datetime(record.ended_at),
                    "is_active": is_active,
                    "overdue_days": self._calculate_overdue_days(
                        overdue_since_date,
                        is_active=is_active,
                        ended_at=record.ended_at,
                        today=today,
                    ),
                    "phase": record.phase,
                    "status_label": record.release.get_status_display() if record.release else "-",
                    "assignees": [],
                }
            )

        return release_records

    def _build_test_plan_records(
        self,
        *,
        slug: str,
        project_ids: List[str],
        today,
        status_filter: str,
    ) -> List[Dict[str, Any]]:
        # TestPlan 当前没有历史逾期记录模型，因此仅返回当前仍逾期的数据
        if status_filter == "resolved":
            return []

        test_plan_queryset = TestPlan.objects.filter(
            project__workspace__slug=slug,
            project__deleted_at__isnull=True,
            project__archived_at__isnull=True,
            deleted_at__isnull=True,
            end_time__isnull=False,
            end_time__lt=today,
        ).exclude(state=TestPlan.State.COMPLETED)

        if project_ids:
            test_plan_queryset = test_plan_queryset.filter(project_id__in=project_ids)

        test_plan_queryset = test_plan_queryset.select_related("project").order_by("-end_time")

        plan_records: List[Dict[str, Any]] = []
        for plan in test_plan_queryset:
            # 逾期从截止日期次日开始计算
            overdue_since_date = plan.end_time + timedelta(days=1) if plan.end_time else None
            plan_records.append(
                {
                    "entity_type": "test_plan",
                    "entity_id": str(plan.id),
                    "name": plan.name,
                    "identifier": None,
                    "project_id": str(plan.project_id) if plan.project_id else None,
                    "project_name": plan.project.name if plan.project else "-",
                    "deadline": self._as_iso_date(plan.end_time),
                    "overdue_since": self._as_iso_date(overdue_since_date),
                    "ended_at": None,
                    "is_active": True,
                    "overdue_days": self._calculate_overdue_days(
                        overdue_since_date,
                        is_active=True,
                        ended_at=None,
                        today=today,
                    ),
                    "phase": None,
                    "status_label": plan.state or "-",
                    "assignees": [],
                }
            )

        return plan_records

    def _collect_records(
        self,
        *,
        slug: str,
        status_filter: str,
        entity_type: Optional[str],
        project_ids: List[str],
    ) -> List[Dict[str, Any]]:
        # 延期分析改为工作区维度统计，不再按当前用户参与项目做裁剪。
        base_filters: Dict[str, Any] = {
            "workspace__slug": slug,
            "project__deleted_at__isnull": True,
            "project__archived_at__isnull": True,
        }
        if project_ids:
            base_filters["project_id__in"] = project_ids
        today = timezone.now().date()

        records: List[Dict[str, Any]] = []
        if entity_type in (None, "issue"):
            records.extend(
                self._build_issue_records(
                    base_filters=base_filters,
                    today=today,
                    status_filter=status_filter,
                )
            )

        if entity_type in (None, "cycle"):
            records.extend(
                self._build_cycle_records(
                    base_filters=base_filters,
                    today=today,
                    status_filter=status_filter,
                )
            )

        if entity_type in (None, "release"):
            records.extend(
                self._build_release_records(
                    base_filters=base_filters,
                    today=today,
                    status_filter=status_filter,
                )
            )

        if entity_type in (None, "test_plan"):
            records.extend(
                self._build_test_plan_records(
                    slug=slug,
                    project_ids=project_ids,
                    today=today,
                    status_filter=status_filter,
                )
            )

        records.sort(
            key=lambda item: (
                1 if item["is_active"] else 0,
                item["overdue_days"],
                item.get("overdue_since") or "",
            ),
            reverse=True,
        )

        return records

    @allow_permission([ROLE.ADMIN, ROLE.MEMBER], level="WORKSPACE")
    def get(self, request, slug):
        status_filter, entity_type, project_ids = self._extract_filters_from_request(request)
        validation_error = self._validate_filters(status_filter=status_filter, entity_type=entity_type)
        if validation_error:
            return validation_error

        records = self._collect_records(
            slug=slug,
            status_filter=status_filter,
            entity_type=entity_type,
            project_ids=project_ids,
        )

        summary = {
            "work_items": sum(1 for item in records if item["entity_type"] == "issue"),
            "cycles": sum(1 for item in records if item["entity_type"] == "cycle"),
            "releases": sum(1 for item in records if item["entity_type"] == "release"),
            "test_plans": sum(1 for item in records if item["entity_type"] == "test_plan"),
        }
        summary["total"] = sum(summary.values())

        trend_by_month = defaultdict(int)
        for item in records:
            overdue_since = item.get("overdue_since")
            if overdue_since:
                trend_by_month[overdue_since[:7]] += 1
        trend = [
            {"month": month, "count": trend_by_month[month]}
            for month in sorted(trend_by_month.keys())
        ]

        return Response(
            {
                "summary": summary,
                "records": records,
                "trend": trend,
            },
            status=status.HTTP_200_OK,
        )


class WorkspaceOverdueAnalyticsExportEndpoint(WorkspaceOverdueAnalyticsEndpoint):
    @staticmethod
    def _format_assignees(record: Dict[str, Any]) -> str:
        assignees = record.get("assignees") or []
        names = [
            assignee.get("display_name")
            for assignee in assignees
            if isinstance(assignee, dict) and assignee.get("display_name")
        ]
        return ", ".join(names) if names else "-"

    @staticmethod
    def _format_date(value: Optional[str]) -> str:
        if not value:
            return "-"
        return str(value).split("T")[0]

    @staticmethod
    def _format_status(record: Dict[str, Any]) -> str:
        status_label = record.get("status_label") or "-"
        return f"仍在延期 · {status_label}" if record.get("is_active") else f"已恢复 · {status_label}"

    @allow_permission([ROLE.ADMIN, ROLE.MEMBER], level="WORKSPACE")
    def get(self, request, slug):
        status_filter, entity_type, project_ids = self._extract_filters_from_request(request)
        date_field = request.GET.get("date_field", "deadline")
        start_date_raw = request.GET.get("start_date")
        end_date_raw = request.GET.get("end_date")

        validation_error = self._validate_filters(status_filter=status_filter, entity_type=entity_type)
        if validation_error:
            return validation_error

        if date_field not in self.ALLOWED_DATE_FIELDS:
            return Response(
                {"error": "date_field 必须是 deadline 或 overdue_since"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        start_date = self._normalize_filter_date(start_date_raw)
        if start_date_raw and not start_date:
            return Response(
                {"error": "start_date 必须是 YYYY-MM-DD 格式"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        end_date = self._normalize_filter_date(end_date_raw)
        if end_date_raw and not end_date:
            return Response(
                {"error": "end_date 必须是 YYYY-MM-DD 格式"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if start_date and end_date and start_date > end_date:
            return Response(
                {"error": "start_date 不能晚于 end_date"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        records = self._collect_records(
            slug=slug,
            status_filter=status_filter,
            entity_type=entity_type,
            project_ids=project_ids,
        )
        records = self._apply_date_filter(
            records,
            date_field=date_field,
            start_date=start_date,
            end_date=end_date,
        )

        columns = [
            ("名称", "name"),
            ("类型", "entity_type"),
            ("项目", "project_name"),
            ("状态", "status_label"),
            ("截止日期", "deadline"),
            ("延期开始", "overdue_since"),
            ("延期天数", "overdue_days"),
            ("负责人", "assignees"),
        ]

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "延期记录"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_align = Alignment(horizontal="center", vertical="center")

        worksheet.append([label for label, _ in columns])
        for col_idx, _ in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for record in records:
            worksheet.append(
                [
                    record.get("name") or "-",
                    self.ENTITY_LABEL_MAP.get(record.get("entity_type"), "-"),
                    record.get("project_name") or "-",
                    self._format_status(record),
                    self._format_date(record.get("deadline")),
                    self._format_date(record.get("overdue_since")),
                    record.get("overdue_days", 0),
                    self._format_assignees(record),
                ]
            )

        widths = [36, 12, 24, 28, 14, 14, 10, 24]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=index).column_letter
            ].width = width

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"overdue-records-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        response = FileResponse(
            output,
            as_attachment=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return response
