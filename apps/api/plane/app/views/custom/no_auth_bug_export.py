import io
import json
from collections import defaultdict
from datetime import timedelta, timezone as dt_timezone
from urllib.parse import quote
from uuid import UUID

from django.db.models import Prefetch
from django.http import FileResponse
from django.utils import timezone
from openpyxl import Workbook
from rest_framework import serializers
from rest_framework.permissions import AllowAny

from plane.app.views.base import BaseAPIView
from plane.db.models import (
    Issue,
    IssueActivity,
    IssueAssignee,
    TypeExtraFieldValue,
    User,
)


class PublicBugReportExportQuerySerializer(serializers.Serializer):
    workspace_slug = serializers.CharField(required=False, default="kfcd")
    start_date = serializers.DateField(required=False)
    end_date = serializers.DateField(required=False, allow_null=True)

    def validate(self, attrs):
        start_date = attrs.get("start_date")
        if start_date is None:
            start_date = timezone.localdate() - timedelta(days=7)
            attrs["start_date"] = start_date

        end_date = attrs.get("end_date")
        if end_date and end_date < start_date:
            raise serializers.ValidationError(
                {"end_date": "结束日期不能早于开始日期。"}
            )
        return attrs


class PublicBugReportExportAPIView(BaseAPIView):
    """
    公开缺陷导出接口（无需鉴权）：
    按 workspace + 时间范围导出缺陷报告为 Excel。
    """

    authentication_classes = []
    permission_classes = [AllowAny]

    bug_type_names = ("缺陷", "缺陷(软件)")
    OPEN_STATE_NAME = "Open"
    TRACE_ACTIVITY_FIELDS = ("state", "assignees")
    TECH_REASON_FIELD_NAME = "技术原因及解决方案"
    BUG_LEVEL_FIELD_NAME = "缺陷级别"
    BUG_REASON_FIELD_NAME = "缺陷原因"
    export_columns = [
        "序号",
        "ID",
        "标题",
        "技术原因及解决方案",
        "项目",
        "创建时间",
        "当前负责人",
        "创建人",
        "产品类型",
        "状态",
        "缺陷级别",
        "缺陷原因",
    ]
    export_widths = {
        "标题": 40,
        "技术原因及解决方案": 80,
        "项目": 30,
        "创建时间": 20,
        "当前负责人": 20,
        "创建人": 16,
        "产品类型": 16,
        "状态": 15,
        "ID": 20,
        "缺陷原因": 30,
    }

    @staticmethod
    def _to_utc8_string(dt_value):
        if not dt_value:
            return ""

        if timezone.is_naive(dt_value):
            dt_value = timezone.make_aware(dt_value, dt_timezone.utc)

        utc8_tz = dt_timezone(timedelta(hours=8))
        return dt_value.astimezone(utc8_tz).strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def _resolve_user_name(user):
        if not user:
            return ""

        display_name = (getattr(user, "display_name", "") or "").strip()
        if display_name:
            return display_name

        email = (getattr(user, "email", "") or "").strip()
        if "@" in email:
            return email.split("@")[0]
        return email

    @staticmethod
    def _build_issue_url(request, issue):
        workspace_slug = (
            issue.workspace.slug if issue.workspace_id and issue.workspace else ""
        )
        project_identifier = (
            issue.project.identifier if issue.project_id and issue.project else ""
        )
        return request.build_absolute_uri(
            f"/{workspace_slug}/browse/{project_identifier}-{issue.sequence_id}/"
        )

    @staticmethod
    def _build_issue_key(issue):
        if issue.project_id and issue.project and issue.sequence_id:
            return f"{issue.project.identifier}-{issue.sequence_id}"
        return str(issue.id)

    @staticmethod
    def _stringify_extra_field_value(raw_value):
        if raw_value is None:
            return ""

        if isinstance(raw_value, list):
            return ",".join(str(item) for item in raw_value if item not in (None, ""))

        if isinstance(raw_value, dict):
            return json.dumps(raw_value, ensure_ascii=False)

        return str(raw_value)

    def _get_tech_reason(self, issue):
        for item in getattr(issue, "tech_reason_values", []):
            value = self._stringify_extra_field_value(item.value)
            if value:
                return value
        return ""

    def _get_bug_level(self, issue):
        for item in getattr(issue, "bug_level_values", []):
            value = self._stringify_extra_field_value(item.value)
            if value:
                return value
        return ""

    def _get_bug_reason(self, issue):
        for item in getattr(issue, "bug_reason_values", []):
            value = self._stringify_extra_field_value(item.value)
            if value:
                return value
        return ""

    @staticmethod
    def _normalize_identifier(raw_identifier):
        if raw_identifier is None:
            return ""

        identifier = str(raw_identifier).strip()
        if not identifier:
            return ""

        try:
            return str(UUID(identifier))
        except (ValueError, TypeError, AttributeError):
            return ""

    def _parse_assignee_snapshot_ids(self, raw_snapshot):
        if not raw_snapshot:
            return []

        values = str(raw_snapshot).split(",")
        assignee_ids = []
        seen = set()

        for value in values:
            normalized_id = self._normalize_identifier(value)
            if not normalized_id or normalized_id in seen:
                continue

            assignee_ids.append(normalized_id)
            seen.add(normalized_id)

        return assignee_ids

    def _apply_assignee_activity(self, assignee_ids, activity):
        old_identifier = self._normalize_identifier(activity.get("old_identifier"))
        new_identifier = self._normalize_identifier(activity.get("new_identifier"))

        # 工作流审批通过后的负责人活动是“整体快照替换”，仅在 new_value 里存 id 串
        if not old_identifier and not new_identifier:
            return self._parse_assignee_snapshot_ids(activity.get("new_value"))

        next_assignee_ids = list(assignee_ids)
        if new_identifier and new_identifier not in next_assignee_ids:
            next_assignee_ids.append(new_identifier)

        if old_identifier and old_identifier in next_assignee_ids:
            next_assignee_ids.remove(old_identifier)

        return next_assignee_ids

    def _get_current_assignee_ids(self, issue):
        assignee_ids = []
        seen = set()
        for link in getattr(issue, "active_issue_assignees", []):
            assignee_id = self._normalize_identifier(getattr(link, "assignee_id", None))
            if not assignee_id or assignee_id in seen:
                continue

            assignee_ids.append(assignee_id)
            seen.add(assignee_id)

        return assignee_ids

    @staticmethod
    def _activity_state_name(activity, key):
        return (activity.get(key) or "").strip()

    def _resolve_open_assignee_ids(self, issue, issue_activities):
        assignee_ids = []
        current_state_name = None
        last_open_assignee_ids = None

        for activity in issue_activities:
            field = activity.get("field")

            if field == "assignees":
                assignee_ids = self._apply_assignee_activity(assignee_ids, activity)
                if current_state_name == self.OPEN_STATE_NAME:
                    last_open_assignee_ids = list(assignee_ids)
                continue

            if field != "state":
                continue

            old_state_name = self._activity_state_name(activity, "old_value")
            new_state_name = self._activity_state_name(activity, "new_value")
            state_before_transition = current_state_name or old_state_name

            if state_before_transition == self.OPEN_STATE_NAME:
                last_open_assignee_ids = list(assignee_ids)

            current_state_name = new_state_name or None
            if current_state_name == self.OPEN_STATE_NAME:
                last_open_assignee_ids = list(assignee_ids)

        issue_state_name = issue.state.name if issue.state_id and issue.state else ""
        if issue_state_name == self.OPEN_STATE_NAME:
            return self._get_current_assignee_ids(issue)

        return last_open_assignee_ids

    def _prepare_open_assignee_context(self, issues):
        if not issues:
            return {}, {}

        activities = (
            IssueActivity.objects.filter(
                issue_id__in=[issue.id for issue in issues],
                field__in=self.TRACE_ACTIVITY_FIELDS,
                deleted_at__isnull=True,
            )
            .values(
                "issue_id",
                "field",
                "old_value",
                "new_value",
                "old_identifier",
                "new_identifier",
                "created_at",
            )
            .order_by("created_at", "id")
        )

        activity_map = defaultdict(list)
        for activity in activities:
            activity_map[str(activity["issue_id"])].append(activity)

        open_assignee_ids_map = {}
        all_open_assignee_ids = set()
        for issue in issues:
            issue_id = str(issue.id)
            issue_activities = activity_map.get(issue_id, [])
            assignee_ids = self._resolve_open_assignee_ids(
                issue=issue, issue_activities=issue_activities
            )
            open_assignee_ids_map[issue_id] = assignee_ids
            if assignee_ids:
                all_open_assignee_ids.update(assignee_ids)

        open_assignee_name_map = {}
        if all_open_assignee_ids:
            users = User.objects.filter(id__in=all_open_assignee_ids).only(
                "id", "display_name", "email"
            )
            open_assignee_name_map = {
                str(user.id): self._resolve_user_name(user) for user in users
            }

        return open_assignee_ids_map, open_assignee_name_map

    def _get_queryset(self, workspace_slug, start_date, end_date):
        queryset = (
            Issue.objects.filter(
                deleted_at__isnull=True,
                type__name__in=self.bug_type_names,
                workspace__slug=workspace_slug,
                created_at__date__gte=start_date,
            )
            .select_related("workspace", "project", "state", "created_by")
            .prefetch_related(
                Prefetch(
                    "issue_assignee",
                    queryset=IssueAssignee.objects.filter(
                        deleted_at__isnull=True
                    ).select_related("assignee"),
                    to_attr="active_issue_assignees",
                ),
                Prefetch(
                    "type_extra_field_values",
                    queryset=TypeExtraFieldValue.objects.filter(
                        deleted_at__isnull=True,
                        extra_field__name=self.TECH_REASON_FIELD_NAME,
                    )
                    .select_related("extra_field")
                    .order_by("-created_at"),
                    to_attr="tech_reason_values",
                ),
                Prefetch(
                    "type_extra_field_values",
                    queryset=TypeExtraFieldValue.objects.filter(
                        deleted_at__isnull=True,
                        extra_field__name=self.BUG_LEVEL_FIELD_NAME,
                    )
                    .select_related("extra_field")
                    .order_by("-created_at"),
                    to_attr="bug_level_values",
                ),
                Prefetch(
                    "type_extra_field_values",
                    queryset=TypeExtraFieldValue.objects.filter(
                        deleted_at__isnull=True,
                        extra_field__name=self.BUG_REASON_FIELD_NAME,
                    )
                    .select_related("extra_field")
                    .order_by("-created_at"),
                    to_attr="bug_reason_values",
                ),
            )
            .order_by("-created_at")
        )

        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)

        return queryset

    def _serialize_issue_rows(
        self,
        request,
        issues,
        open_assignee_ids_map=None,
        open_assignee_name_map=None,
    ):
        open_assignee_ids_map = open_assignee_ids_map or {}
        open_assignee_name_map = open_assignee_name_map or {}
        rows = []
        for index, issue in enumerate(issues, start=1):
            issue_id = str(issue.id)
            open_assignee_ids = open_assignee_ids_map.get(issue_id)
            assignees = [
                open_assignee_name_map.get(assignee_id, "")
                for assignee_id in open_assignee_ids or []
            ]
            assignees = [name for name in assignees if name]
            issue_url = self._build_issue_url(request, issue)

            rows.append(
                {
                    "序号": index,
                    "ID": issue_url,
                    "标题": issue.name or "",
                    "技术原因及解决方案": self._get_tech_reason(issue),
                    "项目": issue.project.name if issue.project_id else "",
                    "创建时间": self._to_utc8_string(issue.created_at),
                    "当前负责人": ",".join(assignees),
                    "创建人": self._resolve_user_name(
                        getattr(issue, "created_by", None)
                    ),
                    "产品类型": issue.project.product_type if issue.project_id else "",
                    "状态": issue.state.name if issue.state_id else "",
                    "缺陷级别": self._get_bug_level(issue),
                    "缺陷原因": self._get_bug_reason(issue),
                }
            )
        return rows

    def _build_workbook(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "缺陷报告"

        worksheet.append(self.export_columns)

        for row in rows:
            worksheet.append([row.get(column, "") for column in self.export_columns])

        for col_idx, header in enumerate(self.export_columns, start=1):
            width = self.export_widths.get(header)
            if width is None:
                max_length = max(
                    len(str(worksheet.cell(row=row_idx, column=col_idx).value or ""))
                    for row_idx in range(1, worksheet.max_row + 1)
                )
                width = min(max(max_length + 2, 10), 50)

            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[column_letter].width = width

        return workbook

    @staticmethod
    def _build_excel_response(workbook):
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"缺陷报告-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        response = FileResponse(
            output,
            as_attachment=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{quote(filename)}"
        )
        return response

    def get(self, request):
        serializer = PublicBugReportExportQuerySerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        query_params = serializer.validated_data

        issues = list(
            self._get_queryset(
                workspace_slug=query_params["workspace_slug"],
                start_date=query_params["start_date"],
                end_date=query_params.get("end_date"),
            )
        )
        open_assignee_ids_map, open_assignee_name_map = (
            self._prepare_open_assignee_context(issues)
        )
        rows = self._serialize_issue_rows(
            request,
            issues,
            open_assignee_ids_map=open_assignee_ids_map,
            open_assignee_name_map=open_assignee_name_map,
        )
        workbook = self._build_workbook(rows)

        return self._build_excel_response(workbook)
