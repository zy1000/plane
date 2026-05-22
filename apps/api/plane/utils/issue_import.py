# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.
"""
工作项 Excel 导入工具。

设计目标：
- 模板字段定义集中在 `IMPORT_FIELD_DEFINITIONS`，前后端共用相同的 field_key。
- `parse_excel` 只负责把 xlsx 读成一组 raw 行，并保留每行的 row_number。
- `validate_rows` / `build_issues` 共享 `_RowResolver`：在校验阶段就把外键全部 resolve 好，
  避免 import 阶段再次走数据库查询，并能在两阶段返回一致的错误信息。
- `build_issues` 在 `transaction.atomic` 内分两轮：第一轮建 Issue + 关联表，第二轮回填父项。
"""

from __future__ import annotations

import datetime
import random
import re
import uuid
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Iterable

from django.db import transaction
from openpyxl import load_workbook

from plane.db.models import (
    Cycle,
    CycleIssue,
    Issue,
    IssueAssignee,
    IssueLabel,
    IssueType,
    Label,
    Module,
    ModuleIssue,
    ProjectMember,
    Release,
    ReleaseIssue,
    State,
)
from plane.db.models.state import StateGroup
from plane.utils.import_export import (
    REQUIREMENT_FORMAT,
    TABLE_FORMAT,
    build_html_table,
)


# ---------------------------------------------------------------------------
# 字段定义
# ---------------------------------------------------------------------------

FIELD_NAME = "name"
FIELD_TYPE = "type"
FIELD_DESCRIPTION = "description"
FIELD_PRIORITY = "priority"
FIELD_ASSIGNEES = "assignees"
FIELD_LABELS = "labels"
FIELD_MODULE = "module"
FIELD_CYCLE = "cycle"
FIELD_RELEASE = "release"
FIELD_START_DATE = "start_date"
FIELD_TARGET_DATE = "target_date"
FIELD_PARENT = "parent"
FIELD_REQUIREMENT_ITEM = "requirement_item"

IGNORE_FIELD = "__ignore__"

IMPORT_FIELD_DEFINITIONS: list[dict[str, Any]] = [
    {"key": FIELD_NAME, "label": "标题", "required": True},
    {"key": FIELD_TYPE, "label": "类型", "required": True},
    {"key": FIELD_DESCRIPTION, "label": "描述", "required": False},
    {"key": FIELD_PRIORITY, "label": "优先级", "required": False},
    {"key": FIELD_ASSIGNEES, "label": "负责人", "required": False},
    {"key": FIELD_LABELS, "label": "标签", "required": False},
    {"key": FIELD_MODULE, "label": "模块", "required": False},
    {"key": FIELD_CYCLE, "label": "迭代", "required": False},
    {"key": FIELD_RELEASE, "label": "发布", "required": False},
    {"key": FIELD_START_DATE, "label": "开始日期", "required": False},
    {"key": FIELD_TARGET_DATE, "label": "截止日期", "required": False},
    {"key": FIELD_PARENT, "label": "父工作项", "required": False},
    {"key": FIELD_REQUIREMENT_ITEM, "label": "需求项（表格列）", "required": False},
]

REQUIRED_FIELDS = {f["key"] for f in IMPORT_FIELD_DEFINITIONS if f["required"]}
ALL_FIELDS = {f["key"] for f in IMPORT_FIELD_DEFINITIONS}

# 允许多列映射的字段白名单。前端 `MULTI_MAP_FIELDS` 须保持同步。
MULTI_MAP_FIELDS: set[str] = {FIELD_REQUIREMENT_ITEM}

# 模板列名 → field_key（用于生成模板与自动猜测）
DEFAULT_COLUMN_TO_FIELD: dict[str, str] = {
    f["label"]: f["key"] for f in IMPORT_FIELD_DEFINITIONS
}

ROW_NUMBER_COLUMN = "#行号"

# 单文件硬限制
MAX_ROWS = 10000

PRIORITY_ALIASES: dict[str, str] = {
    "urgent": "urgent",
    "high": "high",
    "medium": "medium",
    "low": "low",
    "none": "none",
    "紧急": "urgent",
    "高": "high",
    "中": "medium",
    "低": "low",
    "无": "none",
}


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------


@dataclass
class RowResult:
    row_number: int
    title: str = ""
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    resolved: dict[str, Any] = field(default_factory=dict)
    raw: dict[str, Any] = field(default_factory=dict)

    @property
    def passed(self) -> bool:
        return not self.errors

    def to_dict(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "title": self.title,
            "passed": self.passed,
            "errors": list(self.errors),
            "warnings": list(self.warnings),
            "error_reason": "；".join(self.errors),
            "warning_reason": "；".join(self.warnings),
        }


# ---------------------------------------------------------------------------
# Excel 解析
# ---------------------------------------------------------------------------


def _read_headers(file_obj) -> tuple[list[str], bytes]:
    """读取首行作为列名，并返回原始字节用于后续重复解析。"""
    data = file_obj.read() if hasattr(file_obj, "read") else file_obj
    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    worksheet = workbook.active
    if worksheet is None:
        return [], data
    raw_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(raw_iter)
    except StopIteration:
        return [], data
    headers = [(str(cell).strip() if cell is not None else "") for cell in header_row]
    return headers, data


def inspect_file(file_obj) -> dict[str, Any]:
    """
    解析 Excel 头部：返回列名、推荐映射、数据行数。
    用于前端字段映射界面。
    """
    headers, _ = _read_headers(file_obj)
    suggested: dict[str, str] = {}
    for col_name in headers:
        if not col_name or col_name == ROW_NUMBER_COLUMN:
            continue
        field_key = DEFAULT_COLUMN_TO_FIELD.get(col_name)
        if field_key:
            suggested[col_name] = field_key
        else:
            suggested[col_name] = IGNORE_FIELD

    # 单独再解析一次以统计数据行数（避免在 read_only 模式下重复 iter）。
    if hasattr(file_obj, "seek"):
        try:
            file_obj.seek(0)
        except (OSError, ValueError):
            pass
    rows = parse_excel(file_obj)
    return {
        "headers": [h for h in headers if h and h != ROW_NUMBER_COLUMN],
        "suggested_mapping": suggested,
        "row_count": len(rows),
    }


def parse_excel(file_obj) -> list[dict[str, Any]]:
    """
    读取 xlsx 文件首个 sheet，返回行数据列表。
    每个 dict 的 key 为列表头字符串（保留原样），并额外含 `__row_number__`（excel 真实行号 - 表头偏移）。
    跳过全空行；超过 MAX_ROWS 抛 ValueError 由调用方转 422。
    """
    data = file_obj.read() if hasattr(file_obj, "read") else file_obj
    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    worksheet = workbook.active
    if worksheet is None:
        return []

    raw_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(raw_iter)
    except StopIteration:
        return []

    headers = [(str(cell).strip() if cell is not None else "") for cell in header_row]

    rows: list[dict[str, Any]] = []
    # 模板中可能包含「#行号」首列；它仅作为父项引用占位，不参与字段映射。
    has_row_number_column = bool(headers and headers[0] == ROW_NUMBER_COLUMN)
    row_counter = 0
    for excel_row in raw_iter:
        if not any(cell is not None and str(cell).strip() != "" for cell in excel_row):
            continue
        row_counter += 1
        if row_counter > MAX_ROWS:
            raise ValueError(f"单次导入最多支持 {MAX_ROWS} 行，请拆分后再导入")

        row_dict: dict[str, Any] = {}
        for idx, value in enumerate(excel_row):
            if idx >= len(headers):
                break
            key = headers[idx]
            if not key:
                continue
            row_dict[key] = value

        if has_row_number_column:
            row_number_value = row_dict.get(ROW_NUMBER_COLUMN)
            try:
                if row_number_value is None or str(row_number_value).strip() == "":
                    row_dict["__row_number__"] = row_counter
                else:
                    row_dict["__row_number__"] = int(str(row_number_value).strip())
            except (TypeError, ValueError):
                row_dict["__row_number__"] = row_counter
        else:
            row_dict["__row_number__"] = row_counter

        rows.append(row_dict)

    return rows


# ---------------------------------------------------------------------------
# 字段映射校验
# ---------------------------------------------------------------------------


def validate_mapping(mapping: dict[str, str]) -> tuple[bool, str]:
    """
    检查字段映射是否合法：
    - 每个 required 字段必须恰好被映射到一列。
    - 其它字段最多映射到一列，但 `MULTI_MAP_FIELDS` 中的字段允许多列。
    - field_key 必须是已知字段或 `__ignore__`。
    """
    if not isinstance(mapping, dict) or not mapping:
        return False, "缺少字段映射"

    counter: dict[str, int] = {}
    for column, field_key in mapping.items():
        if not isinstance(field_key, str):
            return False, f"列「{column}」映射格式非法"
        if field_key == IGNORE_FIELD:
            continue
        if field_key not in ALL_FIELDS:
            return False, f"列「{column}」映射到未知属性：{field_key}"
        counter[field_key] = counter.get(field_key, 0) + 1

    for required_key in REQUIRED_FIELDS:
        count = counter.get(required_key, 0)
        label = next(
            (f["label"] for f in IMPORT_FIELD_DEFINITIONS if f["key"] == required_key),
            required_key,
        )
        if count == 0:
            return False, f"必填属性「{label}」未配置映射"
        if count > 1 and required_key not in MULTI_MAP_FIELDS:
            return False, f"属性「{label}」映射了多列"

    for key, count in counter.items():
        if count > 1 and key not in MULTI_MAP_FIELDS:
            label = next(
                (f["label"] for f in IMPORT_FIELD_DEFINITIONS if f["key"] == key), key
            )
            return False, f"属性「{label}」映射了多列"

    return True, ""


def _invert_mapping(mapping: dict[str, str]) -> dict[str, str]:
    """field_key → excel 列名（忽略 __ignore__）"""
    return {value: column for column, value in mapping.items() if value != IGNORE_FIELD}


# ---------------------------------------------------------------------------
# 解析器（共享状态）
# ---------------------------------------------------------------------------


class _RowResolver:
    """
    把 Excel 行根据字段映射 resolve 成可写入的 dict。
    校验阶段和导入阶段都用同一份解析逻辑，避免行为漂移。
    """

    def __init__(self, project, user):
        self.project = project
        self.workspace_id = project.workspace_id
        self.project_id = project.id
        self.user = user

        # 缓存：项目内所有 IssueType。type_by_name 以名称小写为键，避免大小写敏感。
        self._issue_types = list(
            IssueType.objects.filter(
                project_id=self.project_id, deleted_at__isnull=True
            )
        )
        self.type_by_name: dict[str, IssueType] = {
            it.name.strip().lower(): it for it in self._issue_types
        }

        # 缓存：每个 type 的 backlog state（按 sequence 升序的第一个）。
        backlog_states = State.objects.filter(
            project_id=self.project_id,
            group=StateGroup.BACKLOG.value,
            deleted_at__isnull=True,
        ).order_by("sequence")
        self.backlog_state_by_type_id: dict[Any, State] = {}
        for st in backlog_states:
            if (
                st.issue_type_id
                and st.issue_type_id not in self.backlog_state_by_type_id
            ):
                self.backlog_state_by_type_id[st.issue_type_id] = st

        # 缓存：模块 / 迭代 / 发布
        self.module_by_name: dict[str, Module] = {
            m.name.strip(): m
            for m in Module.objects.filter(
                project_id=self.project_id, deleted_at__isnull=True
            )
        }
        self.cycle_by_name: dict[str, list[Cycle]] = {}
        for c in Cycle.objects.filter(
            project_id=self.project_id, deleted_at__isnull=True
        ).order_by("created_at"):
            self.cycle_by_name.setdefault(c.name.strip(), []).append(c)
        self.release_by_name: dict[str, Release] = {
            r.name.strip(): r
            for r in Release.objects.filter(
                project_id=self.project_id, deleted_at__isnull=True
            )
        }

        # 缓存：项目成员（仅活跃的 Member/Admin，role>=15），按 display_name 索引。
        self.member_by_display_name: dict[str, list[Any]] = {}
        members = ProjectMember.objects.filter(
            project_id=self.project_id,
            workspace_id=self.workspace_id,
            is_active=True,
            deleted_at__isnull=True,
            role__gte=15,
        ).select_related("member")
        for pm in members:
            user_obj = pm.member
            if not user_obj:
                continue
            name = (user_obj.display_name or "").strip()
            if not name:
                continue
            self.member_by_display_name.setdefault(name, []).append(user_obj)

        # 缓存：项目内现有 Label（仅项目级，与现有 UI 的「项目标签」概念一致）。
        existing_labels = Label.objects.filter(
            workspace_id=self.workspace_id, project_id=self.project_id
        )
        self.label_by_name: dict[str, Label] = {}
        for label in existing_labels:
            self.label_by_name.setdefault(label.name.strip(), label)

        # 缓存：项目内已有 sequence_id（用于父项校验 + 字典查找）。
        self.existing_sequence_ids: set[int] = set(
            Issue.objects.filter(project_id=self.project_id).values_list(
                "sequence_id", flat=True
            )
        )

    # ---- 字段解析 ---------------------------------------------------------

    def resolve(self, raw_row: dict[str, Any], mapping: dict[str, str]) -> RowResult:
        row_number = int(raw_row.get("__row_number__") or 0)
        result = RowResult(row_number=row_number, raw=dict(raw_row))
        inv = _invert_mapping(mapping)

        # name --------------------------------------------------------------
        name_column = inv.get(FIELD_NAME)
        name_value = _clean_text(raw_row.get(name_column)) if name_column else ""
        if not name_value:
            result.errors.append("标题不能为空")
        elif len(name_value) > 255:
            result.errors.append("标题长度不能超过 255")
        result.title = name_value
        result.resolved[FIELD_NAME] = name_value

        # type --------------------------------------------------------------
        type_column = inv.get(FIELD_TYPE)
        type_value = _clean_text(raw_row.get(type_column)) if type_column else ""
        issue_type = None
        if not type_value:
            result.errors.append("类型不能为空")
        else:
            issue_type = self.type_by_name.get(type_value.lower())
            if issue_type is None:
                result.errors.append(f"类型「{type_value}」在项目中不存在")
        result.resolved[FIELD_TYPE] = issue_type

        # state（强制 backlog） ----------------------------------------------
        if issue_type is not None:
            backlog_state = self.backlog_state_by_type_id.get(issue_type.id)
            if backlog_state is None:
                result.errors.append(
                    f"类型「{type_value}」未配置 Backlog 状态，请先到项目设置中初始化状态"
                )
            result.resolved["state"] = backlog_state
        else:
            result.resolved["state"] = None

        # description -------------------------------------------------------
        desc_column = inv.get(FIELD_DESCRIPTION)
        desc_value = _clean_text(raw_row.get(desc_column)) if desc_column else ""
        if desc_value:
            result.resolved[FIELD_DESCRIPTION] = desc_value

        # requirement_item（多列映射，按列名作为表头，单元格按 \n 拆分行） --------
        req_columns = [
            col for col, key in mapping.items() if key == FIELD_REQUIREMENT_ITEM
        ]
        if req_columns:
            table_payload: dict[str, list[str]] = {}
            max_rows = 0
            for col in req_columns:
                cell = raw_row.get(col)
                if cell is None or str(cell).strip() == "":
                    table_payload[col] = []
                else:
                    values = [v.strip() for v in str(cell).split("\n") if v.strip()]
                    table_payload[col] = values
                    if len(values) > max_rows:
                        max_rows = len(values)

            if max_rows > 0:
                # 补齐各列长度（短列以空字符串填充），保证表格行列对齐。
                for col in table_payload:
                    diff = max_rows - len(table_payload[col])
                    if diff > 0:
                        table_payload[col].extend([""] * diff)

                table_rows = build_html_table(table_payload)
                description_text = desc_value or ""
                final_html = (REQUIREMENT_FORMAT + TABLE_FORMAT).format(
                    table_rows=table_rows,
                    description=(
                        _escape_html(description_text) if description_text else ""
                    ),
                    uuid_1=uuid.uuid4(),
                    uuid_2=uuid.uuid4(),
                    uuid_3=uuid.uuid4(),
                    uuid_4=uuid.uuid4(),
                    uuid_5=uuid.uuid4(),
                )
                result.resolved["__description_html__"] = final_html
                stripped_parts = [description_text] if description_text else []
                for col, values in table_payload.items():
                    cleaned_values = [v for v in values if v]
                    if cleaned_values:
                        stripped_parts.append(f"{col}: {', '.join(cleaned_values)}")
                result.resolved["__description_stripped__"] = "\n".join(stripped_parts)

        # priority ----------------------------------------------------------
        priority_column = inv.get(FIELD_PRIORITY)
        priority_value = (
            _clean_text(raw_row.get(priority_column)) if priority_column else ""
        )
        if priority_value:
            normalized = PRIORITY_ALIASES.get(priority_value.lower())
            if normalized is None:
                result.errors.append(
                    f"优先级「{priority_value}」不合法，可选值：urgent/high/medium/low/none"
                )
            else:
                result.resolved[FIELD_PRIORITY] = normalized

        # assignees ---------------------------------------------------------
        assignees_column = inv.get(FIELD_ASSIGNEES)
        if assignees_column:
            raw_value = raw_row.get(assignees_column)
            names = _split_multi(raw_value)
            users: list[Any] = []
            for name in names:
                candidates = self.member_by_display_name.get(name) or []
                if not candidates:
                    result.errors.append(f"负责人「{name}」不在项目成员中")
                    continue
                if len(candidates) > 1:
                    emails = ", ".join(
                        (u.email or u.username or "") for u in candidates
                    )
                    result.errors.append(
                        f"负责人「{name}」在项目中存在多名同名成员（{emails}），请使用唯一展示名"
                    )
                    continue
                users.append(candidates[0])
            result.resolved[FIELD_ASSIGNEES] = users

        # labels ------------------------------------------------------------
        labels_column = inv.get(FIELD_LABELS)
        if labels_column:
            raw_value = raw_row.get(labels_column)
            names = _split_multi(raw_value)
            new_label_names: list[str] = []
            existing_labels: list[Label] = []
            for name in names:
                if len(name) > 255:
                    result.errors.append(f"标签「{name[:30]}…」长度不能超过 255")
                    continue
                label = self.label_by_name.get(name)
                if label:
                    existing_labels.append(label)
                else:
                    new_label_names.append(name)
            if new_label_names:
                result.warnings.append(f"将新建以下标签：{', '.join(new_label_names)}")
            result.resolved[FIELD_LABELS] = {
                "existing": existing_labels,
                "new": new_label_names,
            }

        # module ------------------------------------------------------------
        module_column = inv.get(FIELD_MODULE)
        module_value = _clean_text(raw_row.get(module_column)) if module_column else ""
        if module_value:
            module = self.module_by_name.get(module_value)
            if module is None:
                result.errors.append(f"模块「{module_value}」在项目中不存在")
            else:
                result.resolved[FIELD_MODULE] = module

        # cycle -------------------------------------------------------------
        cycle_column = inv.get(FIELD_CYCLE)
        cycle_value = _clean_text(raw_row.get(cycle_column)) if cycle_column else ""
        if cycle_value:
            cycles = self.cycle_by_name.get(cycle_value) or []
            if not cycles:
                result.errors.append(f"迭代「{cycle_value}」在项目中不存在")
            else:
                if len(cycles) > 1:
                    result.warnings.append(
                        f"项目中存在多个同名迭代「{cycle_value}」，将关联到最早创建的一个"
                    )
                result.resolved[FIELD_CYCLE] = cycles[0]

        # release -----------------------------------------------------------
        release_column = inv.get(FIELD_RELEASE)
        release_value = (
            _clean_text(raw_row.get(release_column)) if release_column else ""
        )
        if release_value:
            release = self.release_by_name.get(release_value)
            if release is None:
                result.errors.append(f"发布「{release_value}」在项目中不存在")
            else:
                result.resolved[FIELD_RELEASE] = release

        # start_date / target_date ------------------------------------------
        for date_field, column_label in (
            (FIELD_START_DATE, "开始日期"),
            (FIELD_TARGET_DATE, "截止日期"),
        ):
            date_column = inv.get(date_field)
            if not date_column:
                continue
            value = raw_row.get(date_column)
            parsed = _parse_date(value)
            if value is None or value == "":
                continue
            if parsed is None:
                result.errors.append(f"{column_label}格式错误，应为 YYYY-MM-DD")
                continue
            result.resolved[date_field] = parsed

        if (
            FIELD_START_DATE in result.resolved
            and FIELD_TARGET_DATE in result.resolved
            and result.resolved[FIELD_START_DATE] > result.resolved[FIELD_TARGET_DATE]
        ):
            result.errors.append("开始日期不能晚于截止日期")

        # parent ------------------------------------------------------------
        parent_column = inv.get(FIELD_PARENT)
        if parent_column:
            parent_raw = _clean_text(raw_row.get(parent_column))
            if parent_raw:
                parent_token = _parse_parent_token(parent_raw)
                if parent_token is None:
                    result.errors.append(
                        f"父工作项「{parent_raw}」格式错误，请填写已存在的 sequence_id 或本批引用 #行号"
                    )
                else:
                    kind, value = parent_token
                    if kind == "existing":
                        if value not in self.existing_sequence_ids:
                            result.errors.append(
                                f"父工作项 sequence_id={value} 在项目中不存在"
                            )
                        else:
                            result.resolved[FIELD_PARENT] = {
                                "kind": "existing",
                                "value": value,
                            }
                    else:  # kind == 'row'
                        result.resolved[FIELD_PARENT] = {"kind": "row", "value": value}

        return result


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _split_multi(value: Any) -> list[str]:
    """按英文/中文逗号切分，去除空项与首尾空格。"""
    if value is None:
        return []
    text = str(value)
    parts = re.split(r"[,，]", text)
    return [p.strip() for p in parts if p and p.strip()]


def _parse_date(value: Any) -> datetime.date | None:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = _clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_parent_token(text: str) -> tuple[str, int] | None:
    """
    支持两种格式：
    - `#123`：表示本批 Excel 中 row_number=123 的行
    - `123` / `Issue-123`：表示已存在的 sequence_id
    """
    cleaned = text.strip()
    if not cleaned:
        return None
    if cleaned.startswith("#"):
        try:
            return "row", int(cleaned[1:].strip())
        except ValueError:
            return None
    match = re.search(r"(\d+)$", cleaned)
    if not match:
        return None
    try:
        return "existing", int(match.group(1))
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# 对外接口
# ---------------------------------------------------------------------------


def validate_rows(
    rows: list[dict[str, Any]],
    mapping: dict[str, str],
    project,
    user,
) -> dict[str, Any]:
    resolver = _RowResolver(project, user)
    results: list[RowResult] = []
    for raw_row in rows:
        result = resolver.resolve(raw_row, mapping)
        results.append(result)

    _post_validate_parent_refs(results)

    passed_count = sum(1 for r in results if r.passed)
    return {
        "total_count": len(results),
        "passed_count": passed_count,
        "all_passed": passed_count == len(results) and len(results) > 0,
        "results": [r.to_dict() for r in results],
    }


def _post_validate_parent_refs(results: list[RowResult]) -> None:
    """父项引用本批 #行号 的二次校验：被引用行必须存在且自身通过。"""
    by_row_number: dict[int, RowResult] = {
        r.row_number: r for r in results if r.row_number
    }

    for r in results:
        parent_ref = r.resolved.get(FIELD_PARENT)
        if not isinstance(parent_ref, dict):
            continue
        if parent_ref.get("kind") != "row":
            continue
        target_row = parent_ref.get("value")
        target = by_row_number.get(target_row)
        if target is None:
            r.errors.append(f"父工作项引用的行号 #{target_row} 在本次导入中不存在")
            continue
        if target.row_number == r.row_number:
            r.errors.append("父工作项不能引用自身")
            continue
        if not target.passed:
            r.errors.append(f"父工作项引用的行号 #{target_row} 自身未通过校验")

    _detect_parent_cycles(results, by_row_number)


def _detect_parent_cycles(
    results: list[RowResult], by_row_number: dict[int, RowResult]
) -> None:
    """简单 DFS 检测本批父引用是否形成环。"""
    parent_map: dict[int, int] = {}
    for r in results:
        parent_ref = r.resolved.get(FIELD_PARENT)
        if isinstance(parent_ref, dict) and parent_ref.get("kind") == "row":
            parent_map[r.row_number] = parent_ref["value"]

    for start_row, _ in parent_map.items():
        visited: set[int] = set()
        cursor = start_row
        while cursor in parent_map:
            if cursor in visited:
                target = by_row_number.get(start_row)
                if target and "存在循环父子引用" not in "".join(target.errors):
                    target.errors.append("存在循环父子引用")
                break
            visited.add(cursor)
            cursor = parent_map[cursor]


def build_issues(
    rows: list[dict[str, Any]],
    mapping: dict[str, str],
    project,
    user,
) -> dict[str, Any]:
    """
    事务化执行导入。返回成功/失败统计；任一行失败即回滚整批。
    """
    resolver = _RowResolver(project, user)
    results: list[RowResult] = []
    for raw_row in rows:
        result = resolver.resolve(raw_row, mapping)
        results.append(result)
    _post_validate_parent_refs(results)

    failed = [r for r in results if not r.passed]
    if failed:
        return {
            "total_count": len(results),
            "success_count": 0,
            "created_ids": [],
            "failed": [
                {
                    "row_number": r.row_number,
                    "title": r.title,
                    "error": "；".join(r.errors),
                }
                for r in failed
            ],
        }

    created_ids: list[str] = []
    label_cache: dict[str, Label] = dict(resolver.label_by_name)

    try:
        with transaction.atomic():
            row_to_issue: dict[int, Issue] = {}

            # 第一轮：建 Issue + 关联（跳过 parent 字段）
            for r in results:
                issue = _create_issue(
                    r,
                    project=project,
                    user=user,
                    label_cache=label_cache,
                )
                row_to_issue[r.row_number] = issue
                created_ids.append(str(issue.id))

            # 第二轮：回填父项
            for r in results:
                parent_ref = r.resolved.get(FIELD_PARENT)
                if not isinstance(parent_ref, dict):
                    continue
                parent_issue = _resolve_parent_issue(
                    parent_ref, row_to_issue, project_id=project.id
                )
                if parent_issue is None:
                    raise ValueError(
                        f"第 {r.row_number} 行父工作项 {parent_ref} 解析失败"
                    )
                issue = row_to_issue[r.row_number]
                issue.parent = parent_issue
                issue.save(update_fields=["parent", "updated_at"])
    except Exception as exc:  # noqa: BLE001 - 顶层兜底，整批回滚
        return {
            "total_count": len(results),
            "success_count": 0,
            "created_ids": [],
            "failed": [
                {
                    "row_number": 0,
                    "title": "",
                    "error": f"导入失败已回滚：{str(exc)}",
                }
            ],
        }

    return {
        "total_count": len(results),
        "success_count": len(created_ids),
        "created_ids": created_ids,
        "failed": [],
    }


def _create_issue(
    result: RowResult,
    project,
    user,
    label_cache: dict[str, Label],
) -> Issue:
    resolved = result.resolved
    issue = Issue(
        name=resolved[FIELD_NAME],
        project=project,
        workspace_id=project.workspace_id,
        type=resolved.get(FIELD_TYPE),
        state=resolved.get("state"),
        priority=resolved.get(FIELD_PRIORITY) or "none",
        start_date=resolved.get(FIELD_START_DATE),
        target_date=resolved.get(FIELD_TARGET_DATE),
        created_by=user,
        updated_by=user,
        is_draft=False,
    )

    if "__description_html__" in resolved:
        issue.description_html = resolved["__description_html__"]
        issue.description_stripped = resolved.get("__description_stripped__") or ""
    elif FIELD_DESCRIPTION in resolved:
        text = resolved[FIELD_DESCRIPTION]
        issue.description_html = f"<p>{_escape_html(text)}</p>"
        issue.description_stripped = text

    issue.save()

    # assignees
    assignees: Iterable = resolved.get(FIELD_ASSIGNEES) or []
    for u in assignees:
        IssueAssignee.objects.create(
            issue=issue,
            assignee=u,
            project=project,
            workspace_id=project.workspace_id,
            created_by=user,
            updated_by=user,
        )

    # labels
    labels_payload = resolved.get(FIELD_LABELS) or {}
    for label in labels_payload.get("existing", []) or []:
        IssueLabel.objects.create(
            issue=issue,
            label=label,
            project=project,
            workspace_id=project.workspace_id,
            created_by=user,
            updated_by=user,
        )
    for label_name in labels_payload.get("new", []) or []:
        label = label_cache.get(label_name)
        if label is None:
            label = Label.objects.create(
                workspace_id=project.workspace_id,
                project=project,
                name=label_name,
                color=_random_color(),
                created_by=user,
                updated_by=user,
            )
            label_cache[label_name] = label
        IssueLabel.objects.create(
            issue=issue,
            label=label,
            project=project,
            workspace_id=project.workspace_id,
            created_by=user,
            updated_by=user,
        )

    if FIELD_MODULE in resolved:
        ModuleIssue.objects.create(
            module=resolved[FIELD_MODULE],
            issue=issue,
            project=project,
            workspace_id=project.workspace_id,
            created_by=user,
            updated_by=user,
        )

    if FIELD_CYCLE in resolved:
        CycleIssue.objects.create(
            cycle=resolved[FIELD_CYCLE],
            issue=issue,
            project=project,
            workspace_id=project.workspace_id,
            created_by=user,
            updated_by=user,
        )

    if FIELD_RELEASE in resolved:
        ReleaseIssue.objects.create(
            release=resolved[FIELD_RELEASE],
            issue=issue,
            project=project,
            workspace_id=project.workspace_id,
            created_by=user,
            updated_by=user,
        )

    return issue


def _resolve_parent_issue(
    parent_ref: dict[str, Any], row_to_issue: dict[int, Issue], project_id
) -> Issue | None:
    kind = parent_ref.get("kind")
    value = parent_ref.get("value")
    if kind == "row":
        return row_to_issue.get(value)
    if kind == "existing":
        return Issue.objects.filter(project_id=project_id, sequence_id=value).first()
    return None


def _escape_html(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("\n", "<br />")
    )


def _random_color() -> str:
    return "#{:06x}".format(random.randint(0, 0xFFFFFF))


# ---------------------------------------------------------------------------
# 模板字段对外暴露（给视图用）
# ---------------------------------------------------------------------------


def list_field_definitions() -> list[dict[str, Any]]:
    """复制一份返回，避免外部修改原始定义。"""
    return [dict(f) for f in IMPORT_FIELD_DEFINITIONS]
