"""需求条目的 Excel 导入 / 导出。

产品需求与标准库条目共用这一份 —— 两者的行结构、字段来源、写入路径完全同构，区别只
在「内置列怎么裁」和「一个 Sheet 还是多个 Sheet」，都被 SheetSpec 吸收掉了。

三条贯穿全文的约定：

1. **列名就是字段名**。自定义列名取 `RequirementField.name`，值却按字段 UUID 存在
   `Requirement.data` 里 —— 导入必须按列名反查字段。这是本模块存在的理由。

2. **表单(form)字段用多级表头 + 纵向合并承载**，布局与前端网格逐格对齐（见
   `requirement-grid.tsx` 的 rowSpan 渲染）：一条需求占 `max(1, 各表单子记录数的最大值)`
   行，非表单列写第一行并纵向合并，表单子列逐行铺。
   导入时靠「编号与标题都为空」判定续行 —— openpyxl 读合并单元格时只有左上角有值、
   其余是 None，天然吻合，不需要额外的分组列，用户手工插一行也照样成立。

3. **附件 / 图片字段不出列**。Excel 里没有它们的表示，塞个文件名回来也还原不成 asset。
   因此更新走的是「合并而非替换」data：只覆盖表里出现过的列，没出现的列（附件、图片、
   以及已停用的字段）保留原值 —— 否则一次导入就把它们全清空了。

4. **模块列写名称路径**（`A/B` = A 下的子模块 B），与库→产品导入的模块映射同一口径。
   模块不是内容，与状态同为旁路轴：新增行随创建载荷挂靠，更新行走 set_requirement_module，
   不 bump version、不受评审中 / 已关闭闸门限制。路径不存在的模块在真正导入时逐级创建，
   校验预览只给提示、不落库。
"""

from __future__ import annotations

import datetime
import re
import zipfile
from dataclasses import dataclass, field as dataclass_field
from html import escape as html_escape, unescape as html_unescape
from io import BytesIO
from typing import Any, Optional
from urllib.parse import quote
from uuid import uuid4

from django.utils.html import strip_tags
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from plane.db.models import (
    RequirementFieldType,
    RequirementItemStatus,
    RequirementModule,
    RequirementPriority,
)
from plane.utils.requirement import (
    CONTENT_BUILTIN_COLUMNS,
    TITLE_MAX_LENGTH,
    builtin_filter_specs,
    builtin_values_from_row,
    field_attr,
    get_requirement_select_mode,
    get_requirement_select_options,
    resolve_builtin_field_layout,
)


# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

#: 单次导入的硬上限。需求条目不是工作项，量级小得多，5000 行足够；超了让用户加筛选，
#: 而不是把一个几分钟的事务塞进 HTTP 请求里。
MAX_ROWS = 5000

MAX_FILE_SIZE = 5 * 1024 * 1024

#: xlsx 是 zip：5 MB 的包解开可能是几百 MB 的 XML，openpyxl 完整模式会把它全部读进内存。
#: 5000 条 × 30 列的真实需求表解开也就十几 MB，48 MB 足够宽，再大就是构造出来的。
MAX_UNCOMPRESSED_SIZE = 48 * 1024 * 1024

#: 工作表里**物理行**的上限（含表单续行与空行）。MAX_ROWS 数的是需求条数，这里防的是
#: 「一个格子格式化到第 100 万行」这类把循环拖死的文件。
MAX_SHEET_ROWS = MAX_ROWS * 10

#: openpyxl 读不了 xls（那是 OLE2 复合文档，不是 zip）。`issue_import` 的后缀白名单里
#: 放了 .xls，实际会在 load_workbook 里抛一个对用户毫无意义的异常 —— 这里不抄那个坑。
ALLOWED_EXTENSIONS = (".xlsx",)

#: 编号列。它既是导出的可读标识，也是导入 upsert 的匹配键。
SEQUENCE_COLUMN_KEY = "__sequence__"
SEQUENCE_COLUMN_LABEL = "编号"

#: 模块列。值是模块的名称路径（见模块头注释第 4 条），列名与前端 requirement_modules.column 一致。
MODULE_COLUMN_KEY = "__module__"
MODULE_COLUMN_LABEL = "模块"
MODULE_PATH_SEPARATOR = "/"
MODULE_NAME_MAX_LENGTH = RequirementModule._meta.get_field("name").max_length

#: 内置列的中文名直接取筛选那份定义，避免两处各写一遍然后慢慢漂移。
_BUILTIN_LABELS = {spec["id"]: spec["name"] for spec in builtin_filter_specs()}

#: 不出列的字段类型。见模块头注释第 3 条。
UNSUPPORTED_FIELD_TYPES = frozenset(
    {RequirementFieldType.ATTACHMENT, RequirementFieldType.IMAGE}
)

#: 多选值在单元格里的分隔符。导出用第一个，导入全收。
MULTI_VALUE_JOINER = "、"
_MULTI_VALUE_SPLIT_RE = re.compile(r"[、,，;；\n\r]+")

#: 同批引用：父项列填 `#12` 表示「本 Sheet 第 12 行那条需求」。沿用 issue_import 的约定。
_IN_BATCH_PARENT_RE = re.compile(r"^#\s*(\d+)$")

#: 导出时给 Sheet 名去重加的后缀，导入时要能剥掉。
_SHEET_SUFFIX_RE = re.compile(r"~(\d+)$")

_ILLEGAL_SHEET_CHARS_RE = re.compile(r"[\[\]:*?/\\]")

#: 把块级标签换成换行，再 strip_tags —— 直接 strip_tags 会把 `<p>a</p><p>b</p>` 粘成 "ab"。
_BLOCK_BOUNDARY_RE = re.compile(
    r"(?i)</\s*(p|div|li|tr|h[1-6]|blockquote)\s*>|<\s*br\s*/?\s*>"
)

_DATE_INPUT_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日")


def _build_choice_aliases(choices):
    """`{值, 中文标签} -> 值` 的双向别名表，两种写法都收。"""
    aliases = {}
    for value, label in choices:
        aliases[str(value).strip().lower()] = value
        aliases[str(label).strip().lower()] = value
    return aliases


#: 优先级的中文写法。模型标签已与 Issue.PRIORITY_CHOICES 对齐成英文（页面上显示的也是
#: 英文），但导入要**宽进**：手工填表的人惯写中文，而且本次对齐之前导出的文件里就是中文，
#: 那些文件必须还能导回来。导出仍然只写模型标签这一种写法。
PRIORITY_CHINESE_ALIASES = {
    "紧急": RequirementPriority.URGENT.value,
    "高": RequirementPriority.HIGH.value,
    "中": RequirementPriority.MEDIUM.value,
    "低": RequirementPriority.LOW.value,
    "无": RequirementPriority.NONE.value,
}

PRIORITY_ALIASES = {
    **_build_choice_aliases(RequirementPriority.choices),
    **PRIORITY_CHINESE_ALIASES,
}
STATUS_ALIASES = _build_choice_aliases(RequirementItemStatus.choices)

PRIORITY_LABELS = {value: label for value, label in RequirementPriority.choices}
STATUS_LABELS = {value: label for value, label in RequirementItemStatus.choices}

BOOLEAN_TRUE = frozenset({"是", "true", "1", "yes", "y", "√", "✓", "t"})
BOOLEAN_FALSE = frozenset({"否", "false", "0", "no", "n", "×", "x", "f"})


class RequirementExcelError(Exception):
    """整份文件级别的错误（读不开、没有可识别的 Sheet、超限）。逐行错误不走这里。"""


# ---------------------------------------------------------------------------
# 文本转换
# ---------------------------------------------------------------------------


def html_to_text(value):
    """富文本 / 描述 → 单元格里的纯文本。有损，且是刻意的：目标是能手工填、能导回。"""
    if not value:
        return ""
    text = _BLOCK_BOUNDARY_RE.sub("\n", str(value))
    text = html_unescape(strip_tags(text))
    lines = [line.strip() for line in text.replace("\r\n", "\n").split("\n")]
    # 连续空行压成一个，否则每个 </p> 都会留下一行空白
    collapsed = []
    for line in lines:
        if not line and (not collapsed or not collapsed[-1]):
            continue
        collapsed.append(line)
    return "\n".join(collapsed).strip()


def text_to_html(value):
    """单元格纯文本 → 富文本。逐行包 `<p>`，内容转义。"""
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    lines = [line.strip() for line in text.replace("\r\n", "\n").split("\n")]
    return "".join(f"<p>{html_escape(line)}</p>" for line in lines if line)


def sanitize_sheet_name(name):
    """按 Excel 的规则收敛 Sheet 名：非法字符去掉、截断到 31 字符、不能为空。"""
    cleaned = _ILLEGAL_SHEET_CHARS_RE.sub("", str(name or "").strip())
    cleaned = cleaned.strip("'")
    if not cleaned:
        cleaned = "需求"
    return cleaned[:31]


def _sheet_name_base(name):
    """剥掉导出时加的去重后缀，让「按单类型导出的文件」也能导回多类型的作用域。"""
    return _SHEET_SUFFIX_RE.sub("", str(name or "").strip())


def normalize_header(value):
    """表头单元格 → 用于匹配的键。全角空格与不间断空格都要清掉，它们肉眼不可见。"""
    if value is None:
        return ""
    return str(value).replace("　", " ").replace("\xa0", " ").strip()


def format_module_path(path):
    """模块名称路径 → 单元格文本。导出渲染与导入的「格子没动」比对共用。"""
    return MODULE_PATH_SEPARATOR.join(path or ())


def parse_module_path(raw):
    """单元格文本 → 模块名称路径元组，返回 (path, error)。空格子返回 ()。

    按 `/` 切分、逐段去首尾空白；空段（`A//B`、`/A`、`A/`）与超长段直接报错，
    而不是静默吞掉 —— 那会让「A/」悄悄变成「A」。
    """
    text = _cell_text(raw)
    if not text:
        return (), None
    segments = [segment.strip() for segment in text.split(MODULE_PATH_SEPARATOR)]
    if any(not segment for segment in segments):
        return None, "模块路径格式不正确：各级名称之间用 / 分隔，且不能为空。"
    if any(len(segment) > MODULE_NAME_MAX_LENGTH for segment in segments):
        return None, f"模块每级名称最长 {MODULE_NAME_MAX_LENGTH} 个字符。"
    return tuple(segments), None


# ---------------------------------------------------------------------------
# 列模型
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ExcelColumn:
    """一个叶子列（真正占据 Excel 一列的东西）。"""

    key: str
    #: sequence | module | builtin | field
    kind: str
    label: str
    spec: Optional[Any] = None
    #: 所属表单字段的 id；非表单子列为 None
    form_id: Optional[str] = None


@dataclass(frozen=True)
class FormGroup:
    """一个表单字段及其子列 —— 对应第一行的跨列表头。"""

    field_id: str
    label: str
    columns: tuple
    #: 表单根字段本身的定义。必填校验要看它（「这个表单至少得有一条子记录」）
    spec: Optional[Any] = None


@dataclass
class SheetSpec:
    requirement_type_id: str
    requirement_type_name: str
    sheet_name: str
    #: 按 Excel 从左到右的叶子列
    columns: list
    #: 表单分组，按列顺序
    form_groups: list
    is_library: bool
    #: 必填、启用、但类型是附件/图片（不出列）的根字段名。新增时这些字段在 Excel 里
    #: 根本没法填，与其让写序列化器吐一句不带字段名的 required，不如提前说清楚
    required_unsupported_labels: list = dataclass_field(default_factory=list)
    _index_by_key: Optional[dict] = None

    @property
    def has_forms(self):
        return bool(self.form_groups)

    @property
    def header_rows(self):
        return 2 if self.has_forms else 1

    @property
    def builtin_keys(self):
        return [c.key for c in self.columns if c.kind == "builtin"]

    @property
    def index_by_key(self):
        """列 key -> 1-based 列号。key 全局唯一（内置列名与字段 UUID 撞不上）。"""
        if self._index_by_key is None:
            self._index_by_key = {
                column.key: index
                for index, column in enumerate(self.columns, start=1)
            }
        return self._index_by_key


def _is_exportable(spec):
    return (
        field_attr(spec, "is_active", True)
        and field_attr(spec, "field_type") not in UNSUPPORTED_FIELD_TYPES
    )


def build_sheet_spec(
    *,
    requirement_type_id,
    requirement_type_name,
    field_specs,
    is_library,
    sheet_name,
    builtin_layout=None,
):
    """把一个需求类型的字段定义摊成 Excel 的列清单。

    列序与前端网格同一套规则：编号、标题、模块锁定最前，随后内置列与自定义根字段按
    builtin_field_layout 的 sort_order 归并（相等时内置在前 —— 旧客户端保存过的
    自定义 sort_order 可能与内置撞值）。builtin_layout 不传即缺省布局（内置在前）。
    """
    columns = [
        ExcelColumn(
            key=SEQUENCE_COLUMN_KEY, kind="sequence", label=SEQUENCE_COLUMN_LABEL
        ),
        # 标题与编号一样锁定在最前，不参与归并
        ExcelColumn(key="title", kind="builtin", label=_BUILTIN_LABELS.get("title", "title")),
        # 模块不是内置内容列（不在 builtin_field_layout 里），库 / 产品两侧都出
        ExcelColumn(key=MODULE_COLUMN_KEY, kind="module", label=MODULE_COLUMN_LABEL),
    ]

    builtin_entries = [
        entry
        for entry in resolve_builtin_field_layout(builtin_layout)
        if not is_library or entry["show_in_library"]
    ]

    specs = list(field_specs)
    children_by_parent = {}
    for spec in specs:
        parent_id = field_attr(spec, "parent_field_id")
        if parent_id:
            children_by_parent.setdefault(str(parent_id), []).append(spec)

    form_groups = []
    required_unsupported = []

    def append_builtin_column(entry):
        column = entry["key"]
        columns.append(
            ExcelColumn(
                key=column, kind="builtin", label=_BUILTIN_LABELS.get(column, column)
            )
        )

    def append_field_columns(spec):
        if not _is_exportable(spec):
            if (
                field_attr(spec, "is_active", True)
                and field_attr(spec, "is_required", False)
                and field_attr(spec, "field_type") in UNSUPPORTED_FIELD_TYPES
            ):
                required_unsupported.append(field_attr(spec, "name") or "")
            return
        field_id = str(field_attr(spec, "id"))
        name = field_attr(spec, "name") or ""
        if field_attr(spec, "field_type") != RequirementFieldType.FORM:
            columns.append(
                ExcelColumn(key=field_id, kind="field", label=name, spec=spec)
            )
            return

        child_columns = tuple(
            ExcelColumn(
                key=str(field_attr(child, "id")),
                kind="field",
                label=field_attr(child, "name") or "",
                spec=child,
                form_id=field_id,
            )
            for child in children_by_parent.get(field_id, [])
            if _is_exportable(child)
        )
        # 一个子列都没有的表单在 Excel 里没有落点（子字段全是附件，或者压根没配）——
        # 整组跳过，而不是留一个空的跨列表头让导入端去猜。
        if not child_columns:
            return
        form_groups.append(
            FormGroup(field_id=field_id, label=name, columns=child_columns, spec=spec)
        )
        columns.extend(child_columns)

    # 双指针归并：builtin_entries 已按 sort_order 升序，根 specs 由调用方按
    # (sort_order, created_at, id) 排好；相等时内置在前
    root_specs = [spec for spec in specs if not field_attr(spec, "parent_field_id")]
    builtin_index = 0
    for spec in root_specs:
        spec_sort_order = field_attr(spec, "sort_order") or 0
        while (
            builtin_index < len(builtin_entries)
            and builtin_entries[builtin_index]["sort_order"] <= spec_sort_order
        ):
            append_builtin_column(builtin_entries[builtin_index])
            builtin_index += 1
        append_field_columns(spec)
    for entry in builtin_entries[builtin_index:]:
        append_builtin_column(entry)

    return SheetSpec(
        requirement_type_id=str(requirement_type_id),
        requirement_type_name=requirement_type_name or "",
        sheet_name=sheet_name,
        columns=columns,
        form_groups=form_groups,
        is_library=is_library,
        required_unsupported_labels=required_unsupported,
    )


def build_sheet_specs(
    *, requirement_types, fields_by_type, is_library, builtin_layout_by_type=None
):
    """requirement_types 是 [(id, name)]，顺序即 Sheet 顺序。

    `fields_by_type` 必须是**已经按作用域筛过**的字段：标准库要传
    `get_library_field_specs(library)` 的结果（它是 show_in_library 规则的唯一执行点），
    产品传 `field_specs_for_requirement_types()` 的分组。这里不再筛一遍 —— 多一个执行点
    就多一处将来会和它对不上的地方。

    `builtin_layout_by_type` 是 {类型 id(str): builtin_field_layout JSON}，缺项走缺省布局。
    """
    used = {}
    specs = []
    for requirement_type_id, name in requirement_types:
        base = sanitize_sheet_name(name)
        count = used.get(base, 0) + 1
        used[base] = count
        # 重名加 ~2 / ~3 后缀，且要留出后缀的位置再截断，否则 31 字符上限会把它切掉
        sheet_name = base if count == 1 else f"{base[: 31 - len(f'~{count}')]}~{count}"
        specs.append(
            build_sheet_spec(
                requirement_type_id=requirement_type_id,
                requirement_type_name=name,
                field_specs=fields_by_type.get(str(requirement_type_id), []),
                is_library=is_library,
                sheet_name=sheet_name,
                builtin_layout=(builtin_layout_by_type or {}).get(
                    str(requirement_type_id)
                ),
            )
        )
    return specs


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------


@dataclass
class ExportContext:
    """导出一次所需的全部旁路数据，一次查好，逐行不再碰库。"""

    #: 编号前缀（ECOM-1 里的 ECOM）。库作用域不参与编号，见 is_library
    scope_identifier: str = ""
    #: user_id(str) -> 展示名
    user_display: dict = dataclass_field(default_factory=dict)
    #: requirement_id(str) -> sequence_id，父项列要拿它拼父需求的编号
    sequence_by_id: dict = dataclass_field(default_factory=dict)
    #: 库作用域：编号是行上手填的 code，不是前缀拼接
    is_library: bool = False
    #: requirement_id(str) -> code，库作用域的父项列用它
    code_by_id: dict = dataclass_field(default_factory=dict)
    #: module_id(str) -> 名称路径文本。行查询没有 select_related("module")（要给
    #: select_for_update 让路），所以必须一次建好索引，不能逐行取 requirement.module
    module_path_by_id: dict = dataclass_field(default_factory=dict)

    def module_text(self, module_id):
        if not module_id:
            return ""
        return self.module_path_by_id.get(str(module_id), "")

    def display_id(self, sequence_id):
        if sequence_id is None:
            return ""
        if not self.scope_identifier:
            return str(sequence_id)
        return f"{self.scope_identifier}-{sequence_id}"

    def row_display_id(self, requirement):
        if self.is_library:
            return requirement.code or ""
        return self.display_id(requirement.sequence_id)

    def display_id_of(self, requirement_id):
        if not requirement_id:
            return ""
        if self.is_library:
            return self.code_by_id.get(str(requirement_id), "")
        return self.display_id(self.sequence_by_id.get(str(requirement_id)))

    def user_name(self, user_id):
        """成员的展示名。

        `user_display` 只装活跃的工作区成员；已经离开工作区的人还挂在旧需求上，这里按需
        补查一次并缓存 —— 导出写空会让人以为负责人被清了，而导入端还要拿同一个名字去判
        「这一格动没动」。
        """
        if not user_id:
            return ""
        key = str(user_id)
        if key not in self.user_display:
            from plane.db.models import User

            found = (
                User.objects.filter(id=key)
                .values_list("display_name", "email")
                .first()
            )
            self.user_display[key] = (
                (found[0] or found[1] or "").strip() if found else ""
            )
        return self.user_display[key]


def _put(worksheet, row, column, value):
    """写一个单元格，顺手堵住 openpyxl 的两个坑：

    - 以 `=` 开头的字符串会被当成**公式**写入 —— 一条标题叫 `=HYPERLINK(...)` 的需求，
      导出的表在 Excel 里打开就会执行。写完强制回字符串类型。
    - 含 \x00-\x1f 控制字符的字符串直接抛 IllegalCharacterError —— 从别处粘贴进来的
      标题里偶尔就有，导出不该因此 500。先剥掉。
    空串写成 None：既省体积，读回来时也不用区分「空串」和「没有」。
    """
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if not value:
            value = None
    cell = worksheet.cell(row=row, column=column, value=value)
    if isinstance(value, str) and cell.data_type != "s":
        cell.data_type = "s"
    return cell


_HEADER_FILL = PatternFill("solid", fgColor="F1F3F5")
_GROUP_FILL = PatternFill("solid", fgColor="E4E8EC")
_HEADER_FONT = Font(bold=True)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def get_form_rows(data, form_id):
    """一个表单字段的子记录列表。形状不对时当空处理 —— 导出不该被脏数据打断。"""
    value = (data or {}).get(form_id)
    if not isinstance(value, list):
        return []
    return [row for row in value if isinstance(row, dict)]


def _max_form_rows(data, sheet_spec):
    return max(
        (len(get_form_rows(data, group.field_id)) for group in sheet_spec.form_groups),
        default=0,
    )


def format_leaf_value(spec, value, ctx):
    """一个自定义字段（或表单子字段）的值 → 单元格内容。"""
    field_type = field_attr(spec, "field_type")

    if field_type == RequirementFieldType.BOOLEAN:
        if value is None:
            return ""
        return "是" if value else "否"

    if field_type == RequirementFieldType.MEMBER:
        return ctx.user_name(value)

    if field_type == RequirementFieldType.RICH_TEXT:
        return html_to_text(value)

    if field_type == RequirementFieldType.SELECT:
        labels = {
            str(option.get("id")): str(option.get("label") or option.get("id") or "")
            for option in get_requirement_select_options(spec)
            if isinstance(option, dict) and option.get("id")
        }
        if get_requirement_select_mode(spec) == "multiple":
            values = value if isinstance(value, list) else []
            return MULTI_VALUE_JOINER.join(
                labels.get(str(item), str(item)) for item in values
            )
        if value in (None, ""):
            return ""
        return labels.get(str(value), str(value))

    if value is None:
        return ""
    return str(value)


def format_builtin_value(column_key, requirement, ctx):
    if column_key == "title":
        return requirement.title or ""
    if column_key == "description_html":
        return html_to_text(requirement.description_html)
    if column_key == "status":
        return STATUS_LABELS.get(requirement.status, requirement.status or "")
    if column_key == "priority":
        return PRIORITY_LABELS.get(requirement.priority, requirement.priority or "")
    if column_key == "assignee_id":
        return ctx.user_name(requirement.assignee_id)
    if column_key in ("start_date", "target_date"):
        return getattr(requirement, column_key, None)
    if column_key == "parent_id":
        return ctx.display_id_of(requirement.parent_id)
    return ""


def _write_header(worksheet, sheet_spec):
    header_rows = sheet_spec.header_rows
    written_forms = set()
    for index, column in enumerate(sheet_spec.columns, start=1):
        if column.form_id is None:
            cell = _put(worksheet, 1, index, column.label)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGNMENT
            if header_rows == 2:
                worksheet.merge_cells(
                    start_row=1, start_column=index, end_row=2, end_column=index
                )
            continue

        # 表单子列：第一次遇到某个表单时写它的跨列分组表头
        if column.form_id not in written_forms:
            written_forms.add(column.form_id)
            group = next(
                item
                for item in sheet_spec.form_groups
                if item.field_id == column.form_id
            )
            group_cell = _put(worksheet, 1, index, group.label)
            group_cell.fill = _GROUP_FILL
            group_cell.font = _HEADER_FONT
            group_cell.alignment = _HEADER_ALIGNMENT
            if len(group.columns) > 1:
                worksheet.merge_cells(
                    start_row=1,
                    start_column=index,
                    end_row=1,
                    end_column=index + len(group.columns) - 1,
                )
        child_cell = _put(worksheet, 2, index, column.label)
        child_cell.fill = _HEADER_FILL
        child_cell.font = _HEADER_FONT
        child_cell.alignment = _HEADER_ALIGNMENT


def _write_requirement(worksheet, sheet_spec, requirement, ctx, start_row):
    """写一条需求，返回它占了几行。"""
    data = requirement.data or {}
    span = max(1, _max_form_rows(data, sheet_spec))

    for index, column in enumerate(sheet_spec.columns, start=1):
        if column.form_id is not None:
            continue
        if column.kind == "sequence":
            value = ctx.row_display_id(requirement)
        elif column.kind == "module":
            value = ctx.module_text(requirement.module_id)
        elif column.kind == "builtin":
            value = format_builtin_value(column.key, requirement, ctx)
        else:
            value = format_leaf_value(column.spec, data.get(column.key), ctx)

        cell = _put(worksheet, start_row, index, value)
        cell.alignment = _BODY_ALIGNMENT
        if isinstance(value, (datetime.date, datetime.datetime)):
            cell.number_format = "yyyy-mm-dd"
        # 纵向合并整组 —— 与网格的 rowSpan 一一对应，也让导入端的「续行」判定成立
        if span > 1:
            worksheet.merge_cells(
                start_row=start_row,
                start_column=index,
                end_row=start_row + span - 1,
                end_column=index,
            )

    for group in sheet_spec.form_groups:
        form_rows = get_form_rows(data, group.field_id)
        for offset in range(span):
            row = form_rows[offset] if offset < len(form_rows) else None
            values = (row or {}).get("values") or {}
            for column in group.columns:
                index = sheet_spec.index_by_key[column.key]
                value = (
                    format_leaf_value(column.spec, values.get(column.key), ctx)
                    if row is not None
                    else ""
                )
                cell = _put(worksheet, start_row + offset, index, value)
                cell.alignment = _BODY_ALIGNMENT

    return span


def _apply_widths(worksheet, sheet_spec):
    for index, column in enumerate(sheet_spec.columns, start=1):
        # 中文按两个字符宽估，再留 4 个字符的余量；上下限防止极端列名把表撑坏
        width = min(max(len(column.label) * 2 + 4, 12), 48)
        if column.key == "description_html":
            width = 48
        worksheet.column_dimensions[get_column_letter(index)].width = width


def write_workbook(sheet_specs, rows_by_type, ctx, *, template=False):
    """把若干 SheetSpec 与它们的行写成一个 xlsx，返回 BytesIO。"""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_spec in sheet_specs:
        worksheet = workbook.create_sheet(title=sheet_spec.sheet_name)
        _write_header(worksheet, sheet_spec)
        cursor = sheet_spec.header_rows + 1
        if not template:
            for requirement in rows_by_type.get(sheet_spec.requirement_type_id, []):
                cursor += _write_requirement(
                    worksheet, sheet_spec, requirement, ctx, cursor
                )
        _apply_widths(worksheet, sheet_spec)
        worksheet.freeze_panes = worksheet.cell(
            row=sheet_spec.header_rows + 1, column=1
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# 解析：xlsx -> 分组后的原始行
# ---------------------------------------------------------------------------


@dataclass
class SheetLayout:
    """一个 Sheet 的表头解析结果。同一个 Sheet 的所有行共用它。"""

    sheet_spec: SheetSpec
    sheet_name: str
    header_rows: int
    #: Excel 列号 -> ExcelColumn
    column_by_index: dict
    #: 表里真实出现过的内置列 key
    present_builtin: set
    #: 表里真实出现过的根字段 id（表单按 form_id 计）
    present_field_ids: set
    #: 没能对上任何列的表头，原样回报给用户
    ignored_headers: list
    #: 表里有没有模块列。没有就整列不动（旧版导出的文件天然如此），不算未知列
    has_module_column: bool = False

    @property
    def index_by_column_key(self):
        """列 key -> **文件里**的列号。不能用 SheetSpec 的列号：用户可能删列或调序。"""
        return {column.key: index for index, column in self.column_by_index.items()}


@dataclass
class ParsedGroup:
    """一条需求在表里占的那一组行（主行 + 若干续行）。"""

    layout: SheetLayout
    row_number: int
    builtin_raw: dict = dataclass_field(default_factory=dict)
    field_raw: dict = dataclass_field(default_factory=dict)
    form_raw: dict = dataclass_field(default_factory=dict)

    @property
    def row_key(self):
        return f"{self.layout.sheet_name}!{self.row_number}"


def _is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _read_header_row(worksheet, row_index, max_column):
    """读一行表头，并把横向合并的分组标题铺满它覆盖的所有列。"""
    values = {
        column: normalize_header(worksheet.cell(row=row_index, column=column).value)
        for column in range(1, max_column + 1)
    }
    for merged in worksheet.merged_cells.ranges:
        if merged.min_row != row_index:
            continue
        anchor = normalize_header(
            worksheet.cell(row=merged.min_row, column=merged.min_col).value
        )
        for column in range(merged.min_col, merged.max_col + 1):
            values[column] = anchor
    return values


def _build_sheet_layout(worksheet, sheet_spec):
    max_column = worksheet.max_column or 0
    if not max_column:
        return None

    first_row = _read_header_row(worksheet, 1, max_column)
    group_labels = {group.label: group for group in sheet_spec.form_groups if group.label}
    # 表头是一行还是两行由**文件**决定而不是由字段定义决定：用户可能删掉了整个表单列组，
    # 那时文件里就只有一行表头，按定义去读第二行会把首条数据当表头吃掉。
    header_rows = 2 if any(value in group_labels for value in first_row.values()) else 1
    second_row = (
        _read_header_row(worksheet, 2, max_column) if header_rows == 2 else {}
    )

    leaf_by_label = {}
    for column in sheet_spec.columns:
        if column.form_id is None and column.label:
            leaf_by_label.setdefault(column.label, column)
    child_by_group = {
        group.label: {
            child.label: child for child in group.columns if child.label
        }
        for group in sheet_spec.form_groups
    }

    # 手工做的两行表头往往不合并单元格，只在分组的第一列写了「验收项」，右边的子列
    # 第一行是空的。遇到「第一行空、第二行有值」就沿用左边最近的分组标题。
    if header_rows == 2:
        last_group = ""
        for index in range(1, max_column + 1):
            top = first_row.get(index, "")
            if top in group_labels:
                last_group = top
            elif top:
                last_group = ""
            elif second_row.get(index, "") and last_group:
                first_row[index] = last_group

    column_by_index = {}
    claimed = set()
    ignored_headers = []
    for index in range(1, max_column + 1):
        top = first_row.get(index, "")
        child = second_row.get(index, "")
        matched = None
        if child and top in child_by_group:
            matched = child_by_group[top].get(child)
        if matched is None and top:
            matched = leaf_by_label.get(top)
        label = top or child
        if matched is None:
            # 认不出来的列一律忽略而不是报错 —— 用户在导出的表上加一列备注是很常见的
            if label:
                ignored_headers.append(label)
            continue
        if matched.key in claimed:
            # 同名列出现两次：第一次出现的赢，后面的当作未知列
            ignored_headers.append(label)
            continue
        claimed.add(matched.key)
        column_by_index[index] = matched

    present_builtin = {
        column.key for column in column_by_index.values() if column.kind == "builtin"
    }
    present_field_ids = set()
    for column in column_by_index.values():
        if column.kind != "field":
            continue
        present_field_ids.add(column.form_id or column.key)

    return SheetLayout(
        sheet_spec=sheet_spec,
        sheet_name=worksheet.title,
        header_rows=header_rows,
        column_by_index=column_by_index,
        present_builtin=present_builtin,
        present_field_ids=present_field_ids,
        ignored_headers=ignored_headers,
        has_module_column=any(
            column.kind == "module" for column in column_by_index.values()
        ),
    )


def _guard_uncompressed_size(content):
    """在交给 openpyxl 之前看一眼解压后的总大小。见 MAX_UNCOMPRESSED_SIZE。"""
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            total = sum(info.file_size for info in archive.infolist())
    except zipfile.BadZipFile as exc:
        raise RequirementExcelError("无法读取 Excel 文件：不是有效的 xlsx。") from exc
    if total > MAX_UNCOMPRESSED_SIZE:
        raise RequirementExcelError(
            f"文件解压后超过 {MAX_UNCOMPRESSED_SIZE // 1024 // 1024} MB，请拆分后再导入。"
        )


def match_sheet_specs(workbook, sheet_specs):
    """Sheet 名 -> SheetSpec。先按精确名字配，配不上再剥掉去重后缀配一次。

    剥后缀这一步是为了让「按单个需求类型筛选后导出的文件」也能导回多类型的作用域 ——
    导出时 Sheet 名是否带 `~2` 取决于当时的类型集合，两次不一定一致。
    """
    exact = {spec.sheet_name: spec for spec in sheet_specs}
    by_base = {}
    for spec in sheet_specs:
        by_base.setdefault(_sheet_name_base(spec.sheet_name), []).append(spec)

    # 只有一个需求类型、文件里也只有一个工作表 —— 没有第二种可能，不必逼用户把
    # 「Sheet1」改名。标准库正是这种情形：自己新建工作簿往里粘也该能导。
    if len(sheet_specs) == 1 and len(workbook.sheetnames) == 1:
        return [(workbook.sheetnames[0], sheet_specs[0])], []

    matched = []
    unmatched = []
    for title in workbook.sheetnames:
        spec = exact.get(title)
        if spec is None:
            candidates = by_base.get(_sheet_name_base(title)) or []
            spec = candidates[0] if len(candidates) == 1 else None
        if spec is None:
            unmatched.append(title)
        else:
            matched.append((title, spec))
    return matched, unmatched


def parse_workbook(file_obj, sheet_specs):
    """xlsx -> (分组后的原始行, 被忽略的 Sheet 名, 被忽略的表头)。

    刻意**不用** `read_only=True`：只读模式拿不到 merged_cells，而纵向合并正是「续行」
    判定与分组表头铺开的依据。
    """
    content = file_obj.read()
    _guard_uncompressed_size(content)
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl 抛的异常类型五花八门
        raise RequirementExcelError(f"无法读取 Excel 文件：{exc}") from exc

    matched, unmatched = match_sheet_specs(workbook, sheet_specs)
    if not matched:
        available = "、".join(spec.sheet_name for spec in sheet_specs)
        raise RequirementExcelError(
            f"没有可识别的工作表。工作表名需要与需求类型同名，当前可用：{available}"
        )

    groups = []
    ignored_headers = []
    for title, sheet_spec in matched:
        worksheet = workbook[title]
        layout = _build_sheet_layout(worksheet, sheet_spec)
        if layout is None:
            continue
        ignored_headers.extend(
            f"{title}!{label}" for label in layout.ignored_headers
        )
        groups.extend(_parse_sheet_rows(worksheet, layout))
        if len(groups) > MAX_ROWS:
            raise RequirementExcelError(
                f"单次导入最多 {MAX_ROWS} 条需求，请先用筛选拆分文件。"
            )

    return groups, unmatched, ignored_headers


def _parse_sheet_rows(worksheet, layout):
    sequence_index = None
    title_index = None
    for index, column in layout.column_by_index.items():
        if column.kind == "sequence":
            sequence_index = index
        elif column.key == "title":
            title_index = index
    child_indexes = layout.index_by_column_key
    mapped_indexes = sorted(layout.column_by_index)
    if not mapped_indexes:
        return []

    max_row = worksheet.max_row or 0
    if max_row - layout.header_rows > MAX_SHEET_ROWS:
        raise RequirementExcelError(
            f"工作表「{worksheet.title}」超过 {MAX_SHEET_ROWS} 行。"
            "请删掉末尾的空行（Ctrl+End 可以定位到最后一个格子），或拆分文件。"
        )

    groups = []
    current = None
    # values_only 的 iter_rows 不会为空位创建 Cell 对象；逐格 ws.cell() 会 —— 一个格式化到
    # 十万行的空表能凭空造出几百万个 Cell
    for offset, values in enumerate(
        worksheet.iter_rows(
            min_row=layout.header_rows + 1,
            max_row=max_row,
            max_col=mapped_indexes[-1],
            values_only=True,
        )
    ):
        row_index = layout.header_rows + 1 + offset
        cells = {index: values[index - 1] for index in mapped_indexes}
        if all(_is_blank(value) for value in cells.values()):
            # 整行皆空：跳过。既不开新组也不断开当前组 —— 表格中间的空行是排版留白，
            # 不该让它改变分组语义
            continue

        starts_group = not _is_blank(cells.get(sequence_index)) or not _is_blank(
            cells.get(title_index)
        )
        if starts_group or current is None:
            current = ParsedGroup(layout=layout, row_number=row_index)
            groups.append(current)
            for index, column in layout.column_by_index.items():
                value = cells.get(index)
                if column.form_id is not None:
                    continue
                if column.kind == "field":
                    current.field_raw[column.key] = value
                else:
                    current.builtin_raw[column.key] = value

        # 表单子记录：每个表单各自收集，整行子列全空就跳过这一条
        for group in layout.sheet_spec.form_groups:
            if group.field_id not in layout.present_field_ids:
                continue
            row_values = {}
            for child in group.columns:
                index = child_indexes.get(child.key)
                if index is None:
                    continue
                row_values[child.key] = cells.get(index)
            if all(_is_blank(value) for value in row_values.values()):
                continue
            current.form_raw.setdefault(group.field_id, []).append(row_values)

    return groups


# ---------------------------------------------------------------------------
# 单元格取值
# ---------------------------------------------------------------------------


def parse_sequence(raw, scope_identifier):
    """`ECOM-12` / `12` -> 12。前缀对不上直接报错。

    前缀校验不是洁癖：从产品 A 导出的表误导进产品 B 时，光看数字会静默更新到一批
    完全无关的需求上，而且是批量的、看不出来的。
    """
    if _is_blank(raw):
        return None, None
    if isinstance(raw, bool):
        return None, "编号格式不正确。"
    if isinstance(raw, (int, float)) and float(raw).is_integer():
        return int(raw), None

    text = str(raw).strip()
    if text.isdigit():
        return int(text), None
    match = re.match(r"^(.*?)[-‑–]\s*(\d+)$", text)
    if not match:
        return None, f"编号「{text}」格式不正确，应形如 {scope_identifier or 'ABC'}-12。"
    prefix, number = match.group(1).strip(), int(match.group(2))
    if scope_identifier and prefix.upper() != scope_identifier.upper():
        return None, (
            f"编号「{text}」不属于当前范围，应以 {scope_identifier}- 开头。"
        )
    return number, None


def parse_date_value(raw):
    if _is_blank(raw):
        return None, None
    if isinstance(raw, datetime.datetime):
        return raw.date(), None
    if isinstance(raw, datetime.date):
        return raw, None
    text = str(raw).strip()
    for pattern in _DATE_INPUT_FORMATS:
        try:
            return datetime.datetime.strptime(text, pattern).date(), None
        except ValueError:
            continue
    return None, f"日期「{text}」无法识别，请使用 2026-08-18 这样的格式。"


def parse_boolean_value(raw):
    if _is_blank(raw):
        return None, None
    if isinstance(raw, bool):
        return raw, None
    text = str(raw).strip().lower()
    if text in BOOLEAN_TRUE:
        return True, None
    if text in BOOLEAN_FALSE:
        return False, None
    return None, f"「{raw}」不是有效的是/否值。"


def _select_option_index(spec):
    """label(小写去空格) / id 都能查到 option id。"""
    index = {}
    labels = []
    for option in get_requirement_select_options(spec):
        if not isinstance(option, dict) or not option.get("id"):
            continue
        option_id = str(option["id"])
        label = str(option.get("label") or "")
        index[option_id.strip().lower()] = option_id
        if label:
            index[label.strip().lower()] = option_id
            labels.append(label)
    return index, labels


def parse_select_value(spec, raw):
    index, labels = _select_option_index(spec)
    available = "、".join(labels) or "（该字段还没有配置选项）"
    multiple = get_requirement_select_mode(spec) == "multiple"

    if _is_blank(raw):
        return ([] if multiple else None), None

    tokens = [
        token.strip()
        for token in _MULTI_VALUE_SPLIT_RE.split(str(raw))
        if token.strip()
    ]
    if not multiple:
        if len(tokens) > 1:
            return None, f"该字段是单选，但填了多个值。可选：{available}"
        option_id = index.get(tokens[0].lower()) if tokens else None
        if option_id is None:
            return None, f"选项「{tokens[0]}」不存在。可选：{available}"
        return option_id, None

    resolved = []
    for token in tokens:
        option_id = index.get(token.lower())
        if option_id is None:
            return None, f"选项「{token}」不存在。可选：{available}"
        if option_id not in resolved:
            resolved.append(option_id)
    return resolved, None


# ---------------------------------------------------------------------------
# 逐行解析结果
# ---------------------------------------------------------------------------


@dataclass
class RowResult:
    sheet: str
    row_number: int
    row_key: str
    client_id: str
    requirement_type_id: str
    requirement_type_name: str
    title: str = ""
    display_id: str = ""
    #: create | update | unchanged | skip
    #:   update    —— 内容有变化，或只改了状态（内容不动、只写状态轴）
    #:   unchanged —— 与现有行完全一致，什么都不做（导出再导回时绝大多数行都是它）
    #:   skip      —— 命中只读闸门（评审中 / 已关闭），整行不动
    action: str = "create"
    skip_reason: str = ""
    errors: list = dataclass_field(default_factory=list)
    warnings: list = dataclass_field(default_factory=list)
    requirement_id: Optional[str] = None
    version: Optional[int] = None
    #: 库作用域：编号列的手填文本。create 行随载荷落库；update 行它只是匹配键
    code: Optional[str] = None
    builtin: dict = dataclass_field(default_factory=dict)
    data: dict = dataclass_field(default_factory=dict)
    #: 父项指向本批里的另一行时记在这里，等落库拿到新 id 再回填
    parent_row_key: Optional[str] = None
    #: 需要写的交付状态；None 表示这一行不动状态
    status_value: Optional[str] = None
    current_status: Optional[str] = None
    #: 目标模块的名称路径：None = 这一行不动模块；() = 新增不挂 / 更新移回「全部」；
    #: 非空 = 挂到该路径（不存在的在导入时逐级创建）
    module_path: Optional[tuple] = None
    #: 更新行：模块与现有行不同。与 status_value 同为旁路轴，不算内容
    module_changed: bool = False
    #: 更新行：内容（内置内容列 + data）与现有行是否有实质差异。False 时不进批量写入
    content_changed: bool = True
    #: 更新行对应的现有行对象，只在解析期用，不进响应
    existing: Optional[Any] = dataclass_field(default=None, repr=False)

    @property
    def passed(self):
        return not self.errors and self.action in ("create", "update")

    @property
    def writes_content(self):
        """要不要进批量写入（creates / updates）。状态只改状态轴，不算。"""
        return self.passed and (self.action == "create" or self.content_changed)

    def to_payload(self):
        return {
            "sheet": self.sheet,
            "row_number": self.row_number,
            "row_key": self.row_key,
            "title": self.title,
            "display_id": self.display_id,
            "requirement_type_name": self.requirement_type_name,
            "action": self.action,
            "skip_reason": self.skip_reason,
            "passed": self.passed,
            "errors": list(self.errors),
            "warnings": list(self.warnings),
        }


def _norm_scalar(value):
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if value is None or value == "" or value == []:
        return None
    if isinstance(value, str):
        return value
    return str(value) if not isinstance(value, (bool, int, float, list, dict)) else value


def _norm_data(data):
    """把 data 收敛成可比较的形状：空值归 None 并剔除，表单子记录只比 values 不比 id。"""
    normalized = {}
    for key, value in (data or {}).items():
        if isinstance(value, list) and value and all(
            isinstance(item, dict) and "values" in item for item in value
        ):
            rows = [_norm_data(item.get("values") or {}) for item in value]
            normalized[key] = rows
            continue
        scalar = _norm_scalar(value)
        if scalar is None:
            continue
        normalized[key] = scalar
    return normalized


def _cell_text(value):
    """单元格 -> 与导出表示比较用的文本。日期按 ISO，其它 str 后去首尾空白。"""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


class RequirementImportResolver:
    """把一批 ParsedGroup 解析成可直接喂给写序列化器的载荷。

    所有外键在这里一次性 resolve 完 —— 校验端点与导入端点共用同一次解析，两边的
    报错因此不可能不一致（`issue_import` 的 `_RowResolver` 也是这个思路）。

    **更新行的原则：格子没动，值就不动。** 每一格先和「现有值导出后长什么样」比一下，
    一样就沿用现有值 —— 富文本因此不会在导出再导回时被拍扁成 `<p>`，已离开工作区的
    负责人不会被清空，表单子记录的 id 也得以保留。整行都没动的，最后标成 unchanged，
    连 version 都不 bump。
    """

    def __init__(
        self,
        *,
        scope_identifier,
        existing_rows,
        members,
        is_library,
        module_index=None,
    ):
        self.scope_identifier = scope_identifier or ""
        self.is_library = is_library

        # 作用域内的模块树（module_path_index 的产物）。id_by_text 让「名称里本身带 /」
        # 的已有模块也能按整格文本精确命中，不被切分规则误伤
        path_by_id, id_by_path = module_index or ({}, {})
        self.module_path_by_id = dict(path_by_id)
        self.module_id_by_path = dict(id_by_path)
        self.module_id_by_text = {
            format_module_path(path): module_id
            for path, module_id in self.module_id_by_path.items()
        }

        self.rows_by_sequence = {}
        self.rows_by_id = {}
        # 库作用域按手填编号匹配已有行；活行的 code 由
        # req_unique_library_code_active 保证唯一，不会碰撞
        self.rows_by_code = {}
        for row in existing_rows:
            self.rows_by_id[str(row.id)] = row
            if row.sequence_id is not None:
                self.rows_by_sequence[row.sequence_id] = row
            if is_library and row.code:
                self.rows_by_code[row.code] = row

        self.user_display = {}
        self._by_email = {}
        self._by_name = {}
        for user_id, display_name, email in members:
            key = str(user_id)
            name = (display_name or email or "").strip()
            self.user_display[key] = name
            if email:
                self._by_email[email.strip().lower()] = key
            if name:
                self._by_name.setdefault(name.lower(), []).append(key)

        # 「现有值导出后长什么样」要用与导出完全相同的渲染，所以直接造一个 ExportContext
        self.export_ctx = ExportContext(
            scope_identifier=self.scope_identifier,
            user_display=self.user_display,
            sequence_by_id={
                str(row.id): row.sequence_id for row in existing_rows
            },
            is_library=is_library,
            code_by_id={str(row.id): row.code or "" for row in existing_rows},
            module_path_by_id={
                module_id: format_module_path(path)
                for module_id, path in self.module_path_by_id.items()
            },
        )

    # -- 成员 ---------------------------------------------------------------

    def resolve_member(self, raw):
        if _is_blank(raw):
            return None, None
        text = str(raw).strip()
        user_id = self._by_email.get(text.lower())
        if user_id:
            return user_id, None
        candidates = self._by_name.get(text.lower()) or []
        if len(candidates) == 1:
            return candidates[0], None
        if len(candidates) > 1:
            return None, f"工作区里有多个成员叫「{text}」，请改填邮箱。"
        return None, f"成员「{text}」不在本工作区，或已停用。"

    # -- 单个字段 -----------------------------------------------------------

    @staticmethod
    def _is_empty_value(spec, value):
        """判空口径与 validate_requirement_leaf_value 保持一致（含富文本剥标签）。"""
        if field_attr(spec, "field_type") == RequirementFieldType.RICH_TEXT and isinstance(
            value, str
        ):
            return not strip_tags(value).strip()
        return value is None or value == "" or value == []

    def check_required(self, spec, value, label):
        """必填校验，返回错误消息或 None。

        写序列化器本来就会拦必填，但它抛的是不带字段名的 `This field is required.` ——
        `validate_requirement_data` 调 `validate_requirement_leaf_value` 时没有把 field_key
        包进去。一行一个「某个字段必填」对着几百行的预览毫无用处，所以这里提前用列名再判
        一次；序列化器那道仍然留着兜没出现在表里的字段。
        """
        if not field_attr(spec, "is_required", False) or not field_attr(spec, "is_active", True):
            return None
        if self._is_empty_value(spec, value):
            return f"「{label}」不能为空。"
        return None

    def _same_as_export(self, spec, raw, existing_value):
        """这一格是不是「导出时长这样、现在还长这样」。"""
        return _cell_text(raw) == _cell_text(
            format_leaf_value(spec, existing_value, self.export_ctx)
        )

    def resolve_leaf(self, spec, raw, existing_value=None, *, has_existing=False):
        """一个叶子字段的值。has_existing 时先做「没动就沿用」判定。"""
        field_type = field_attr(spec, "field_type")
        # 富文本与成员是有损的两类：富文本导出成纯文本，成员按名字导出（离开工作区的
        # 人名字都拿不到）。格子没动就原样保留，别把已有的东西冲掉
        if has_existing and field_type in (
            RequirementFieldType.RICH_TEXT,
            RequirementFieldType.MEMBER,
        ):
            if self._same_as_export(spec, raw, existing_value):
                return existing_value, None
        if field_type == RequirementFieldType.RICH_TEXT:
            return text_to_html(raw), None
        if field_type == RequirementFieldType.MEMBER:
            return self.resolve_member(raw)
        if field_type == RequirementFieldType.BOOLEAN:
            return parse_boolean_value(raw)
        if field_type == RequirementFieldType.SELECT:
            return parse_select_value(spec, raw)
        if _is_blank(raw):
            return "", None
        return str(raw).strip(), None

    # -- 一整条需求 ---------------------------------------------------------

    def display_id(self, sequence_id):
        return self.export_ctx.display_id(sequence_id)

    def resolve_group(self, group):
        """第一遍：逐格解析。跨行的判定（重复编号、同批父项、成环、闸门）在 finalize。"""
        layout = group.layout
        sheet_spec = layout.sheet_spec
        result = RowResult(
            sheet=layout.sheet_name,
            row_number=group.row_number,
            row_key=group.row_key,
            client_id=str(uuid4()),
            requirement_type_id=sheet_spec.requirement_type_id,
            requirement_type_name=sheet_spec.requirement_type_name,
        )
        existing = self._resolve_target(group, result)
        result.existing = existing
        self._resolve_title(group, result, existing)
        self._resolve_builtin_columns(group, result, existing)
        self._resolve_status(group, result, existing)
        self._resolve_module(group, result, existing)
        self._resolve_parent(group, result, existing)
        self._resolve_data(group, result, existing)
        if existing is None:
            self._check_create_requirements(group, result)
        return result

    def finalize(self, result):
        """第二遍：内容有没有变、能不能写。必须在同批父项回填之后调 —— 父项也是内容。"""
        existing = result.existing
        if existing is None:
            return
        result.content_changed = bool(result.parent_row_key) or self._content_differs(
            result, existing
        )
        if result.errors:
            return

        if (
            not result.content_changed
            and result.status_value is None
            and not result.module_changed
        ):
            result.action = "unchanged"
            return

        if existing.pending_change_item_id:
            # 评审中：内容只读，但状态轴 / 模块轴与评审轴正交 —— 只改状态或模块照样放行
            if result.content_changed:
                pending_type = getattr(existing, "pending_change_type", None)
                result.action = "skip"
                result.status_value = None
                self._drop_module_change(result)
                result.skip_reason = (
                    "删除待审批中，内容只读。"
                    if pending_type == "delete"
                    else "评审中，内容只读。"
                )
            return

        closed = RequirementItemStatus.CLOSED
        if (
            existing.status == closed
            and result.content_changed
            and result.status_value in (None, closed)
        ):
            result.action = "skip"
            result.status_value = None
            self._drop_module_change(result)
            result.skip_reason = "已关闭，内容只读。把「状态」列改成其它值即可重开并写入。"

    @staticmethod
    def _drop_module_change(result):
        """整行被闸门跳过时，模块改动一并作废 —— 跳过就是整行不动。"""
        result.module_path = None
        result.module_changed = False

    def _content_differs(self, result, existing):
        before = builtin_values_from_row(existing)
        after = {**before, **result.builtin}
        for column in CONTENT_BUILTIN_COLUMNS:
            if _norm_scalar(before.get(column)) != _norm_scalar(after.get(column)):
                return True
        return _norm_data(existing.data) != _norm_data(result.data)

    def _resolve_target(self, group, result):
        """按编号找到要更新的行；产品作用域编号为空即新增，库作用域编号必填。"""
        if self.is_library:
            return self._resolve_library_target(group, result)
        sequence_id, error = parse_sequence(
            group.builtin_raw.get(SEQUENCE_COLUMN_KEY), self.scope_identifier
        )
        if error:
            result.errors.append(error)
            return None
        if sequence_id is None:
            return None

        result.display_id = self.display_id(sequence_id)
        existing = self.rows_by_sequence.get(sequence_id)
        if existing is None:
            # 不静默改成新增：手滑填错一个编号就凭空多出一批需求，且没人会发现
            result.errors.append(
                f"编号 {result.display_id} 在当前范围内不存在。要新增请把编号留空。"
            )
            return None

        result.action = "update"
        result.requirement_id = str(existing.id)
        result.version = existing.version
        result.current_status = existing.status
        if str(existing.requirement_type_id) != result.requirement_type_id:
            result.errors.append(
                "该需求属于其它需求类型，而需求类型创建后不可更换。"
                "请把这一行放回它自己的工作表。"
            )
        return existing

    def _resolve_library_target(self, group, result):
        """库作用域按手填编号精确匹配：命中 → 更新该条，未命中 → 新增（编号即该文本）。

        编号是手填文本，不走 parse_sequence（没有前缀规则，**不校验格式**）；
        但必填非空 —— 库条目创建必须带编号，这里没有「留空即新增」。
        编号列对更新行只是匹配键：改一条条目的编号只能在页面上做，不能靠 Excel。
        """
        text = _cell_text(group.builtin_raw.get(SEQUENCE_COLUMN_KEY))
        if not text:
            result.errors.append("编号不能为空。")
            return None
        result.display_id = text
        result.code = text
        existing = self.rows_by_code.get(text)
        if existing is None:
            return None

        result.action = "update"
        result.requirement_id = str(existing.id)
        result.version = existing.version
        result.current_status = existing.status
        return existing

    def _resolve_title(self, group, result, existing):
        layout = group.layout
        if "title" not in layout.present_builtin:
            if existing is not None:
                result.title = existing.title or ""
            else:
                result.errors.append("缺少「标题」列，无法新增需求。")
            return
        raw = group.builtin_raw.get("title")
        title = "" if _is_blank(raw) else str(raw).strip()
        result.title = title
        if not title:
            # 表里出现了标题列却留空，几乎一定是误操作 —— 清空一条需求的标题不是批量导入
            # 该干的事
            result.errors.append("标题不能为空。")
        elif len(title) > TITLE_MAX_LENGTH:
            result.errors.append(f"标题最长 {TITLE_MAX_LENGTH} 个字符。")
        else:
            result.builtin["title"] = title

    def _resolve_builtin_columns(self, group, result, existing):
        """除标题 / 状态 / 父项之外的内置列（那三个各有各的规矩）。"""
        layout = group.layout
        for key in ("description_html", "priority", "assignee_id", "start_date", "target_date"):
            if key not in layout.present_builtin:
                continue
            raw = group.builtin_raw.get(key)
            label = _BUILTIN_LABELS.get(key, key)

            if key == "description_html":
                # 描述是富文本，导出有损：格子没动就保留原 HTML
                if existing is not None and _cell_text(raw) == _cell_text(
                    html_to_text(existing.description_html)
                ):
                    result.builtin[key] = existing.description_html
                else:
                    result.builtin[key] = text_to_html(raw)
            elif key == "priority":
                if _is_blank(raw):
                    result.builtin[key] = RequirementPriority.NONE.value
                    continue
                value = PRIORITY_ALIASES.get(str(raw).strip().lower())
                if value is None:
                    result.errors.append(
                        f"「{label}」「{raw}」不是有效值。可选："
                        + "、".join(PRIORITY_LABELS.values())
                    )
                else:
                    result.builtin[key] = value
            elif key == "assignee_id":
                # 负责人按名字导出，离开工作区的人连名字都拿不到（导出为空）。格子没动
                # 就保留原值，否则一次往返就把这些行的负责人清空了
                if existing is not None and _cell_text(raw) == _cell_text(
                    self.export_ctx.user_name(existing.assignee_id)
                ):
                    result.builtin[key] = (
                        str(existing.assignee_id) if existing.assignee_id else None
                    )
                    continue
                value, error = self.resolve_member(raw)
                if error:
                    result.errors.append(f"「{label}」{error}")
                else:
                    result.builtin[key] = value
            else:
                value, error = parse_date_value(raw)
                if error:
                    result.errors.append(f"「{label}」{error}")
                else:
                    result.builtin[key] = value

    def _resolve_status(self, group, result, existing):
        """状态是交付状态轴，不算内容：新增恒「未开始」，更新走独立写入口。"""
        layout = group.layout
        if self.is_library or "status" not in layout.present_builtin:
            return
        raw = group.builtin_raw.get("status")
        if _is_blank(raw):
            return
        value = STATUS_ALIASES.get(str(raw).strip().lower())
        if value is None:
            result.errors.append(
                f"「状态」「{raw}」不是有效值。可选："
                + "、".join(STATUS_LABELS.values())
            )
            return
        if existing is None:
            if value != RequirementItemStatus.NOT_STARTED:
                result.warnings.append("新增的需求恒为「未开始」，状态列已忽略。")
            return
        if value != existing.status:
            result.status_value = value

    def _resolve_module(self, group, result, existing):
        """模块是旁路轴，不算内容：新增随创建载荷挂靠，更新走 set_requirement_module。

        表里没有模块列就整列不动。空格子：新增不挂；更新且现有行挂着模块 → 移回「全部」
        （与父项留空置 None 同款）。非空先按整格文本精确匹配已有模块，再按 `/` 切分。
        """
        layout = group.layout
        if not layout.has_module_column:
            return
        raw = group.builtin_raw.get(MODULE_COLUMN_KEY)
        existing_module_id = (
            str(existing.module_id) if existing is not None and existing.module_id else None
        )

        if _is_blank(raw):
            if existing is None:
                result.module_path = ()
            elif existing_module_id:
                result.module_path = ()
                result.module_changed = True
            return

        text = _cell_text(raw)
        # 格子没动就不动：与导出渲染的现有路径比对（软删模块渲染为空，比不上就算改了）
        if existing is not None and text == self.export_ctx.module_text(existing_module_id):
            return

        matched_id = self.module_id_by_text.get(text)
        if matched_id is not None:
            path = self.module_path_by_id[matched_id]
        else:
            path, error = parse_module_path(raw)
            if error:
                result.errors.append(f"「{MODULE_COLUMN_LABEL}」{error}")
                return
            if path not in self.module_id_by_path:
                result.warnings.append(
                    f"模块「{format_module_path(path)}」不存在，导入时将自动创建。"
                )

        result.module_path = path
        if existing is not None:
            result.module_changed = True

    def _resolve_parent(self, group, result, existing):
        layout = group.layout
        if "parent_id" not in layout.present_builtin:
            return
        raw = group.builtin_raw.get("parent_id")
        if _is_blank(raw):
            result.builtin["parent_id"] = None
            return

        text = str(raw).strip()
        in_batch = _IN_BATCH_PARENT_RE.match(text)
        if in_batch:
            # `#12` = 本工作表第 12 行那条需求。父项可能是同批新增的、还没有编号的行，
            # 所以留到落库拿到 id 之后再回填
            result.parent_row_key = f"{layout.sheet_name}!{int(in_batch.group(1))}"
            return

        # 格子没动就沿用，省一次解析，也不会因为父项已关闭而被「新指派」规则误伤
        # （库作用域走 display_id_of 的 code 分支，比较口径自动一致）
        if existing is not None and existing.parent_id and text == self.export_ctx.display_id_of(
            existing.parent_id
        ):
            result.builtin["parent_id"] = str(existing.parent_id)
            return

        if self.is_library:
            # 手填编号精确匹配，不走 parse_sequence
            parent = self.rows_by_code.get(text)
        else:
            sequence_id, error = parse_sequence(text, self.scope_identifier)
            if error:
                result.errors.append(f"「父项」{error}")
                return
            parent = self.rows_by_sequence.get(sequence_id)
        if parent is None:
            result.errors.append(f"「父项」编号 {text} 在当前范围内不存在。")
        elif existing is not None and str(parent.id) == str(existing.id):
            result.errors.append("「父项」不能是它自己。")
        else:
            result.builtin["parent_id"] = str(parent.id)

    def _resolve_data(self, group, result, existing):
        """自定义字段。更新时**合并而非替换** —— 见模块头注释第 3 条。"""
        layout = group.layout
        sheet_spec = layout.sheet_spec
        existing_data = (existing.data or {}) if existing is not None else {}
        has_existing = existing is not None
        values = {}

        for column in sheet_spec.columns:
            if column.kind != "field" or column.form_id is not None:
                continue
            if column.key not in layout.present_field_ids:
                continue
            value, error = self.resolve_leaf(
                column.spec,
                group.field_raw.get(column.key),
                existing_data.get(column.key),
                has_existing=has_existing,
            )
            if error:
                result.errors.append(f"「{column.label}」{error}")
                continue
            values[column.key] = value
            required_error = self.check_required(column.spec, value, column.label)
            if required_error:
                result.errors.append(required_error)

        for form_group in sheet_spec.form_groups:
            if form_group.field_id not in layout.present_field_ids:
                continue
            existing_rows = get_form_rows(existing_data, form_group.field_id)
            rows = []
            for index, raw_row in enumerate(group.form_raw.get(form_group.field_id, [])):
                # 按位置对上现有子记录：id 沿用、表里没有的子字段（附件等）沿用
                base = existing_rows[index] if index < len(existing_rows) else None
                row_values = dict((base or {}).get("values") or {})
                for child in form_group.columns:
                    if child.key not in raw_row:
                        continue
                    label = f"{form_group.label} / {child.label}"
                    value, error = self.resolve_leaf(
                        child.spec,
                        raw_row.get(child.key),
                        row_values.get(child.key),
                        has_existing=base is not None,
                    )
                    if error:
                        result.errors.append(f"「{label}」{error}")
                        continue
                    row_values[child.key] = value
                    required_error = self.check_required(child.spec, value, label)
                    if required_error:
                        result.errors.append(required_error)
                rows.append(
                    {
                        "id": (base or {}).get("id") or str(uuid4()),
                        "values": row_values,
                    }
                )
            values[form_group.field_id] = rows
            if not rows:
                required_error = self.check_required(
                    form_group.spec, [], form_group.label
                )
                if required_error:
                    result.errors.append(required_error)

        result.data = {**existing_data, **values}

    def _check_create_requirements(self, group, result):
        """新增行：必填字段必须在表里有对应的列，且能通过 Excel 填。"""
        layout = group.layout
        sheet_spec = layout.sheet_spec
        missing = []
        for column in sheet_spec.columns:
            if column.kind != "field" or column.form_id is not None:
                continue
            if column.key in layout.present_field_ids:
                continue
            if field_attr(column.spec, "is_required", False):
                missing.append(column.label)
        for form_group in sheet_spec.form_groups:
            if form_group.field_id in layout.present_field_ids:
                continue
            if field_attr(form_group.spec, "is_required", False):
                missing.append(form_group.label)
        if missing:
            result.errors.append(
                "表里缺少必填列：" + "、".join(f"「{name}」" for name in missing)
            )
        if sheet_spec.required_unsupported_labels:
            result.errors.append(
                "字段 "
                + "、".join(f"「{name}」" for name in sheet_spec.required_unsupported_labels)
                + " 为必填的附件/图片，无法通过 Excel 新增，请先在页面上录入。"
            )


# ---------------------------------------------------------------------------
# 跨行校验与批量载荷
# ---------------------------------------------------------------------------


def _flag_duplicate_targets(results):
    """两行指向同一条需求 —— 批量保存不接受重复 id，先在这里给出可读的报错。"""
    first_by_id = {}
    for result in results:
        if result.requirement_id is None:
            continue
        first = first_by_id.get(result.requirement_id)
        if first is None:
            first_by_id[result.requirement_id] = result
        else:
            result.errors.append(
                f"编号 {result.display_id} 在文件里出现了多次"
                f"（第 {first.row_number} 行已经用过）。"
            )


def _flag_duplicate_codes(results):
    """库作用域：两行**新增**用了同一个手填编号 —— 批内撞号在落库前给出可读报错。

    两行匹配到同一条已有条目（update 撞 update）由 _flag_duplicate_targets 拦；
    「新增的编号与库内已有条目重复」不会发生 —— 重复就已经匹配成 update 了。
    产品作用域没有 code，天然空转。
    """
    first_by_code = {}
    for result in results:
        if result.code is None or result.requirement_id is not None:
            continue
        first = first_by_code.get(result.code)
        if first is None:
            first_by_code[result.code] = result
        else:
            result.errors.append(
                f"编号 {result.code} 在文件里出现了多次"
                f"（第 {first.row_number} 行已经用过）。"
            )


def _resolve_in_batch_parents(results, by_key):
    for result in results:
        if not result.parent_row_key:
            continue
        target = by_key.get(result.parent_row_key)
        row_number = result.parent_row_key.rsplit("!", 1)[-1]
        if target is None:
            result.errors.append(f"「父项」引用的第 {row_number} 行不存在。")
        elif target is result:
            result.errors.append("「父项」不能是它自己。")
        elif target.requirement_id:
            # 目标是已存在的需求，直接落 id，不需要等落库
            result.builtin["parent_id"] = target.requirement_id
            result.parent_row_key = None
        elif target.errors:
            result.errors.append(
                f"「父项」引用的第 {target.row_number} 行没有通过校验。"
            )


def _detect_parent_cycles(results, by_key):
    for result in results:
        seen = set()
        cursor = result
        while cursor is not None and cursor.parent_row_key:
            if cursor.row_key in seen:
                result.errors.append("「父项」在本次导入的行之间构成了循环。")
                break
            seen.add(cursor.row_key)
            cursor = by_key.get(cursor.parent_row_key)


def resolve_groups(groups, resolver):
    """把解析出的分组变成逐行结果。校验端点与导入端点共用这一个入口。"""
    results = [resolver.resolve_group(group) for group in groups]
    by_key = {result.row_key: result for result in results}
    _flag_duplicate_targets(results)
    _flag_duplicate_codes(results)
    _resolve_in_batch_parents(results, by_key)
    _detect_parent_cycles(results, by_key)
    for result in results:
        resolver.finalize(result)
    return results


def summarize(results, *, ignored_sheets=None, ignored_headers=None):
    """校验端点的响应体。"""
    return {
        "total_count": len(results),
        "create_count": sum(1 for r in results if r.passed and r.action == "create"),
        "update_count": sum(1 for r in results if r.passed and r.action == "update"),
        "unchanged_count": sum(1 for r in results if r.action == "unchanged"),
        "skipped_count": sum(1 for r in results if r.action == "skip"),
        "error_count": sum(1 for r in results if r.errors),
        "all_passed": bool(results) and not any(r.errors for r in results),
        "ignored_sheets": list(ignored_sheets or []),
        "ignored_headers": list(ignored_headers or []),
        "results": [r.to_payload() for r in results],
    }


def build_batch_payload(results, *, selected_keys=None):
    """选中的、通过校验的行 -> (选中行, creates, updates, parent_by_client_id)。

    只改状态的更新行在 `chosen` 里但**不在** `updates` 里 —— 状态不算内容，走独立写入口。
    `parent_by_client_id` 的形状与 `utils.requirement.remap_imported_parents` 完全一致，
    落库后直接喂给它回填同批父子关系。
    """
    chosen = [
        result
        for result in results
        if result.passed
        and (selected_keys is None or result.row_key in selected_keys)
    ]
    client_by_row_key = {result.row_key: result.client_id for result in chosen}

    creates = []
    updates = []
    parent_by_client_id = {}
    for result in chosen:
        if result.parent_row_key:
            parent_client_id = client_by_row_key.get(result.parent_row_key)
            if parent_client_id:
                parent_by_client_id[result.client_id] = parent_client_id
        if result.action == "create":
            create_item = {
                "client_id": result.client_id,
                "data": result.data,
                "builtin": result.builtin,
                "requirement_type_id": result.requirement_type_id,
            }
            # 库作用域：新增行带手填编号；update 行不带 —— 编号列只是匹配键
            if result.code is not None:
                create_item["code"] = result.code
            creates.append(create_item)
        elif result.writes_content:
            updates.append(
                {
                    "id": result.requirement_id,
                    "data": result.data,
                    "builtin": result.builtin,
                    "version": result.version,
                }
            )
    return chosen, creates, updates, parent_by_client_id


def collect_module_paths(chosen):
    """选中行里要挂靠的全部模块路径（去重、非空），供导入前逐级取或建。"""
    return {
        result.module_path for result in chosen if result.module_path
    }


def assign_create_modules(creates, chosen, module_id_by_path):
    """把已创建好的模块 id 按 client_id **原位**写进 create 载荷。

    不重建列表：`_excel_batch_error` 靠 creates 的下标把序列化器报错映射回行。
    """
    path_by_client_id = {
        result.client_id: result.module_path
        for result in chosen
        if result.action == "create" and result.module_path
    }
    for item in creates:
        path = path_by_client_id.get(item["client_id"])
        if path:
            item["module_id"] = module_id_by_path[path]


def module_update_groups(chosen, module_id_by_path):
    """更新行的模块改动按目标模块分组：{module_id | None: [requirement_id]}。

    None 一组是「移回全部」。每组一次 set_requirement_module，与批量移动同一个写入口。
    """
    groups = {}
    for result in chosen:
        if result.action != "update" or not result.module_changed:
            continue
        target = module_id_by_path[result.module_path] if result.module_path else None
        groups.setdefault(target, []).append(result.requirement_id)
    return groups


def split_status_changes(chosen):
    """状态改动分两拨，顺序是刻意的。

    重开（目标态非「已关闭」）必须在写内容**之前**做，否则已关闭的行会被写闸门挡下，
    「把状态改回进行中并同时改标题」这种最自然的用法就只生效一半；而关闭必须在写内容
    **之后**做，否则自己刚设的 closed 会把自己这一行的内容更新挡住。
    """
    closed = RequirementItemStatus.CLOSED
    before = [r for r in chosen if r.status_value and r.status_value != closed]
    after = [r for r in chosen if r.status_value == closed]
    return before, after


def validate_upload(file_obj):
    """上传文件的门槛校验，返回错误消息或 None。"""
    if file_obj is None:
        return "未上传文件。"
    if file_obj.size and file_obj.size > MAX_FILE_SIZE:
        return f"文件大小不能超过 {MAX_FILE_SIZE // 1024 // 1024} MB。"
    if not str(file_obj.name or "").lower().endswith(ALLOWED_EXTENSIONS):
        return "仅支持 .xlsx 文件，请另存为 xlsx 后重试。"
    return None


def attach_download_filename(response, filename):
    """带中文文件名的下载头。与 `views/issue/base.py` 里那份同形 —— 刻意重写一份而不是
    跨模块 import，那个模块拖着整个工作项视图的依赖链。"""
    response["Content-Disposition"] = f"attachment; filename*=UTF-8\'\'{quote(filename)}"
    exposed = {
        item.strip()
        for item in (response.get("Access-Control-Expose-Headers") or "").split(",")
        if item.strip()
    }
    exposed.add("Content-Disposition")
    response["Access-Control-Expose-Headers"] = ", ".join(sorted(exposed))
    return response


def workspace_member_rows(workspace_id):
    """`[(user_id, display_name, email)]` —— 导出渲染负责人、导入反查负责人共用一份。

    口径与 `validate_requirement_builtin_values` 里的成员校验一致（活跃的工作区成员），
    否则会出现「导入时说这个人不在工作区，但导出的表里就是他」这种自相矛盾。
    """
    from plane.db.models import WorkspaceMember

    return list(
        WorkspaceMember.objects.filter(
            workspace_id=workspace_id, is_active=True, deleted_at__isnull=True
        )
        .select_related("member")
        .values_list("member_id", "member__display_name", "member__email")
    )


def flatten_serializer_errors(errors, sheet_spec=None, prefix=""):
    """DRF 的嵌套错误 -> 人能读的一维列表，并尽量把字段 UUID 换成列名。

    校验预览必须和真正落库跑同一套校验，否则「预览全绿、导入报错」——所以这里要能把
    写序列化器吐出来的结构摊平回具体某一列。
    """
    labels = {}
    if sheet_spec is not None:
        for column in sheet_spec.columns:
            labels[column.key] = column.label
        for group in sheet_spec.form_groups:
            labels[group.field_id] = group.label
    labels.update(_BUILTIN_LABELS)
    labels["module_id"] = MODULE_COLUMN_LABEL

    messages = []
    if isinstance(errors, dict):
        for key, value in errors.items():
            label = labels.get(str(key))
            if key in ("builtin", "data", "non_field_errors"):
                head = prefix
            else:
                head = f"{prefix}「{label or key}」"
            messages.extend(flatten_serializer_errors(value, sheet_spec, head))
    elif isinstance(errors, (list, tuple)):
        for item in errors:
            messages.extend(flatten_serializer_errors(item, sheet_spec, prefix))
    else:
        text = str(errors).strip()
        if text:
            messages.append(f"{prefix}{text}" if prefix else text)
    return messages
